
# ============================================================
# UPSTOX NIFTY 5-MIN HISTORICAL OPTION CHAIN
# ============================================================
#
# Python 3.14
#
# INSTALL:
#
# pip install requests pandas numpy openpyxl
#
# TOKEN:
#
# C:\Users\soiku\Desktop\access_token.txt
#
# OUTPUT:
#
# C:\Users\soiku\Desktop\Upstox_NIFTY_5Min_OptionChain.xlsx
#
# ============================================================
#
# IMPORTANT OI LOGIC
#
# CE_OI_RAW       = Upstox historical raw OI
# PE_OI_RAW       = Upstox historical raw OI
#
# CE_OI           = CE_OI_RAW / 65
# PE_OI           = PE_OI_RAW / 65
#
# CE_Change_OI    = daily CE_OI - previous trading day CE_OI
# PE_Change_OI    = daily PE_OI - previous trading day PE_OI
#
# PCR_OI          = PE_OI / CE_OI
#
# IMPORTANT VOLUME LOGIC
#
# CE_Volume_RAW   = Upstox 5-minute candle volume
# PE_Volume_RAW   = Upstox 5-minute candle volume
#
# CE_Daily_Volume_RAW
#                 = sum of 5-minute raw volumes for that
#                   option on that trading date
#
# CE_Daily_Volume
#                 = CE_Daily_Volume_RAW / 65
#
# Example:
#
# 355007705 / 65 = 5461657
#
# ============================================================

import os
import math
import time
import traceback
import sqlite3

from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import requests
import numpy as np
import pandas as pd


# ============================================================
# SETTINGS
# ============================================================

BASE_DIR = r"C:\Users\soiku\Desktop"

TOKEN_FILE = os.path.join(
    BASE_DIR,
    "access_token.txt"
)

OUTPUT_FILE = os.path.join(
    BASE_DIR,
    "Upstox_NIFTY_5Min_OptionChain.xlsx"
)

DB_FILE = os.path.join(
    BASE_DIR,
    "Upstox_NIFTY_OptionChain.db"
)


# ============================================================
# UPSTOX API
# ============================================================

API_V2 = "https://api.upstox.com/v2"

API_V3 = "https://api.upstox.com/v3"


# ============================================================
# NIFTY
# ============================================================

UNDERLYING_KEY = "NSE_INDEX|Nifty 50"

UNDERLYING_NAME = "NIFTY"


# ============================================================
# EXPIRY
# ============================================================

EXPIRY_DATE = "2026-09-01"


# ============================================================
# HISTORICAL RANGE
# ============================================================

START_DATE = "2026-08-01"

END_DATE = "2026-09-01"


# ============================================================
# 5 MINUTE
# ============================================================

INTERVAL_UNIT = "minutes"

INTERVAL = "5"


# ============================================================
# LOT SIZE
# ============================================================

LOT_SIZE = 65


# ============================================================
# STRIKES
#
# ATM +/- 10 strikes
# ============================================================

STRIKES_EACH_SIDE = 10

STRIKE_STEP = 50


# ============================================================
# IV
# ============================================================

RISK_FREE_RATE = 0.10

DAYS_PER_YEAR = 365.0


# ============================================================
# TIMEZONE
# ============================================================

IST = ZoneInfo(
    "Asia/Kolkata"
)


# ============================================================
# DISPLAY
# ============================================================

pd.set_option(
    "display.max_columns",
    None
)

pd.set_option(
    "display.width",
    300
)

pd.set_option(
    "display.max_rows",
    100
)


# ============================================================
# HEADER
# ============================================================

print()
print("=" * 80)
print("UPSTOX NIFTY 5-MIN HISTORICAL OPTION CHAIN")
print("=" * 80)
print()

print(
    "Underlying :",
    UNDERLYING_KEY
)

print(
    "Expiry     :",
    EXPIRY_DATE
)

print(
    "History    :",
    START_DATE,
    "->",
    END_DATE
)

print(
    "Interval   :",
    "5 minutes"
)

print(
    "Lot Size   :",
    LOT_SIZE
)

print()

print(
    "Output     :",
    OUTPUT_FILE
)

print(
    "Database   :",
    DB_FILE
)

print()


# ============================================================
# LOAD TOKEN
# ============================================================

print("=" * 80)
print("LOADING ACCESS TOKEN")
print("=" * 80)
print()


if not os.path.exists(TOKEN_FILE):

    print("ERROR:")
    print(
        "access_token.txt was not found."
    )

    print()
    print(
        "Create:"
    )

    print(
        TOKEN_FILE
    )

    raise SystemExit


with open(
    TOKEN_FILE,
    "r",
    encoding="utf-8"
) as f:

    ACCESS_TOKEN = f.read().strip()


if not ACCESS_TOKEN:

    print("ERROR:")
    print(
        "access_token.txt is empty."
    )

    raise SystemExit


print(
    "Access token loaded."
)

print()


# ============================================================
# HTTP SESSION
# ============================================================

session = requests.Session()

session.headers.update(
    {
        "Authorization":
            f"Bearer {ACCESS_TOKEN}",

        "Accept":
            "application/json",

        "Content-Type":
            "application/json"
    }
)


# ============================================================
# API GET
# ============================================================

def api_get(
    url,
    params=None,
    timeout=30
):

    try:

        response = session.get(
            url,
            params=params,
            timeout=timeout
        )

    except Exception as error:

        print()
        print(
            "NETWORK ERROR"
        )

        print(
            error
        )

        return None


    if response.status_code != 200:

        print()
        print(
            "UPSTOX API ERROR"
        )

        print(
            "HTTP:",
            response.status_code
        )

        print(
            "URL:",
            response.url
        )

        try:

            print(
                response.json()
            )

        except Exception:

            print(
                response.text[:1000]
            )

        return None


    try:

        return response.json()

    except Exception as error:

        print()
        print(
            "JSON ERROR"
        )

        print(
            error
        )

        return None


# ============================================================
# ALL-STRIKE OPTION CHAIN
# ============================================================
# Upstox provides a dedicated option-chain endpoint containing
# every strike for the requested expiry, with CE/PE market data
# and Greeks.
# ============================================================

def get_option_chain(expiry_date):

    url = API_V2 + "/option/chain"

    data = api_get(
        url,
        params={
            "instrument_key": UNDERLYING_KEY,
            "expiry_date": expiry_date
        }
    )

    if not data:
        return []

    try:
        rows = data.get("data", [])
        return rows if isinstance(rows, list) else []
    except Exception:
        return []


def build_all_strike_chain(expiry_date):

    raw_rows = get_option_chain(expiry_date)
    records = []

    for row in raw_rows:

        try:
            strike = safe_float(row.get("strike_price"))
            spot = safe_float(row.get("underlying_spot_price"))
            expiry = row.get("expiry", expiry_date)
            pcr = safe_float(row.get("pcr"))

            if not np.isfinite(strike):
                continue

            call = row.get("call_options") or {}
            put = row.get("put_options") or {}

            cm = call.get("market_data") or {}
            pm = put.get("market_data") or {}
            cg = call.get("option_greeks") or {}
            pg = put.get("option_greeks") or {}

            call_oi_raw = safe_float(cm.get("oi"))
            put_oi_raw = safe_float(pm.get("oi"))
            call_prev_oi_raw = safe_float(cm.get("prev_oi"))
            put_prev_oi_raw = safe_float(pm.get("prev_oi"))

            records.append({
                "expiry": expiry,
                "underlying_spot": spot,
                "strike": strike,
                "Moneyness": get_moneyness(strike, spot) if np.isfinite(spot) else "",

                "CE_Instrument_Key": call.get("instrument_key"),
                "CE_LTP": safe_float(cm.get("ltp")),
                "CE_Close": safe_float(cm.get("close_price")),
                "CE_Volume_RAW": safe_float(cm.get("volume")),
                "CE_Volume": safe_float(cm.get("volume")) / LOT_SIZE,
                "CE_OI_RAW": call_oi_raw,
                "CE_OI": call_oi_raw / LOT_SIZE if np.isfinite(call_oi_raw) else np.nan,
                "CE_Previous_OI_RAW": call_prev_oi_raw,
                "CE_Previous_OI": call_prev_oi_raw / LOT_SIZE if np.isfinite(call_prev_oi_raw) else np.nan,
                "CE_Change_OI": (call_oi_raw - call_prev_oi_raw) / LOT_SIZE if np.isfinite(call_oi_raw) and np.isfinite(call_prev_oi_raw) else np.nan,
                "CE_Bid": safe_float(cm.get("bid_price")),
                "CE_Bid_Qty": safe_float(cm.get("bid_qty")),
                "CE_Ask": safe_float(cm.get("ask_price")),
                "CE_Ask_Qty": safe_float(cm.get("ask_qty")),
                "CE_IV": safe_float(cg.get("iv")),
                "CE_Delta": safe_float(cg.get("delta")),
                "CE_Gamma": safe_float(cg.get("gamma")),
                "CE_Theta": safe_float(cg.get("theta")),
                "CE_Vega": safe_float(cg.get("vega")),
                "CE_POP": safe_float(cg.get("pop")),

                "PCR_OI": pcr,

                "PE_Instrument_Key": put.get("instrument_key"),
                "PE_LTP": safe_float(pm.get("ltp")),
                "PE_Close": safe_float(pm.get("close_price")),
                "PE_Volume_RAW": safe_float(pm.get("volume")),
                "PE_Volume": safe_float(pm.get("volume")) / LOT_SIZE,
                "PE_OI_RAW": put_oi_raw,
                "PE_OI": put_oi_raw / LOT_SIZE if np.isfinite(put_oi_raw) else np.nan,
                "PE_Previous_OI_RAW": put_prev_oi_raw,
                "PE_Previous_OI": put_prev_oi_raw / LOT_SIZE if np.isfinite(put_prev_oi_raw) else np.nan,
                "PE_Change_OI": (put_oi_raw - put_prev_oi_raw) / LOT_SIZE if np.isfinite(put_oi_raw) and np.isfinite(put_prev_oi_raw) else np.nan,
                "PE_Bid": safe_float(pm.get("bid_price")),
                "PE_Bid_Qty": safe_float(pm.get("bid_qty")),
                "PE_Ask": safe_float(pm.get("ask_price")),
                "PE_Ask_Qty": safe_float(pm.get("ask_qty")),
                "PE_IV": safe_float(pg.get("iv")),
                "PE_Delta": safe_float(pg.get("delta")),
                "PE_Gamma": safe_float(pg.get("gamma")),
                "PE_Theta": safe_float(pg.get("theta")),
                "PE_Vega": safe_float(pg.get("vega")),
                "PE_POP": safe_float(pg.get("pop"))
            })

        except Exception:
            continue

    df = pd.DataFrame(records)

    if not df.empty:
        df = df.sort_values("strike").reset_index(drop=True)
        df["Total_OI"] = df["CE_OI"].fillna(0) + df["PE_OI"].fillna(0)
        df["OI_Difference"] = df["PE_OI"] - df["CE_OI"]
        df["PCR_Change_OI"] = np.where(
            df["CE_Change_OI"].notna() & (df["CE_Change_OI"].abs() > 0),
            df["PE_Change_OI"] / df["CE_Change_OI"],
            np.nan
        )

    return df


# ============================================================
# EXPIRY LIST
# ============================================================

def get_active_expiries():

    data = api_get(
        API_V2 + "/option/contract",
        params={"instrument_key": UNDERLYING_KEY}
    )

    expiries = set()

    if data:
        for item in data.get("data", []):
            expiry = item.get("expiry")
            if expiry:
                expiries.add(str(expiry))

    return sorted(expiries)


def get_past_expiries():

    data = api_get(
        API_V2 + "/expired-instruments/expiries",
        params={"instrument_key": UNDERLYING_KEY}
    )

    if not data:
        return []

    values = data.get("data", [])
    return sorted({str(x) for x in values if x})


def build_expiry_dataframe():

    today_date = datetime.now(IST).date()

    active = get_active_expiries()
    past = get_past_expiries()

    rows = []
    all_dates = sorted(set(active + past))

    for expiry in all_dates:
        try:
            expiry_date = datetime.strptime(expiry, "%Y-%m-%d").date()
        except Exception:
            continue

        if expiry_date < today_date:
            status = "PAST"
            source = "Upstox expired-instruments"
        elif expiry_date == today_date:
            status = "TODAY"
            source = "Upstox active contracts"
        else:
            status = "COMING"
            source = "Upstox active contracts"

        rows.append({
            "Expiry": expiry,
            "Status": status,
            "Source": source
        })

    df = pd.DataFrame(rows)

    if not df.empty:
        df["Expiry"] = pd.to_datetime(df["Expiry"]).dt.date.astype(str)
        df = df.sort_values("Expiry").reset_index(drop=True)

    return df, active, past


# ============================================================
# SAFE FLOAT
# ============================================================

def safe_float(value):

    try:

        if value is None:

            return np.nan

        return float(value)

    except Exception:

        return np.nan


# ============================================================
# NORMAL CDF
# ============================================================

def normal_cdf(x):

    return (
        0.5
        *
        (
            1.0
            +
            math.erf(
                x / math.sqrt(2.0)
            )
        )
    )


# ============================================================
# NORMAL PDF
# ============================================================

def normal_pdf(x):

    return (
        math.exp(
            -0.5 * x * x
        )
        /
        math.sqrt(
            2.0 * math.pi
        )
    )


# ============================================================
# BLACK-SCHOLES CALL
# ============================================================

def bs_call_price(
    S,
    K,
    T,
    r,
    sigma
):

    if (
        S <= 0
        or K <= 0
        or T <= 0
        or sigma <= 0
    ):

        return np.nan


    try:

        d1 = (
            math.log(S / K)
            +
            (
                r
                +
                0.5 * sigma * sigma
            )
            * T
        ) / (
            sigma
            * math.sqrt(T)
        )

        d2 = (
            d1
            -
            sigma
            * math.sqrt(T)
        )

        return (
            S * normal_cdf(d1)
            -
            K
            * math.exp(-r * T)
            * normal_cdf(d2)
        )

    except Exception:

        return np.nan


# ============================================================
# BLACK-SCHOLES PUT
# ============================================================

def bs_put_price(
    S,
    K,
    T,
    r,
    sigma
):

    if (
        S <= 0
        or K <= 0
        or T <= 0
        or sigma <= 0
    ):

        return np.nan


    try:

        d1 = (
            math.log(S / K)
            +
            (
                r
                +
                0.5 * sigma * sigma
            )
            * T
        ) / (
            sigma
            * math.sqrt(T)
        )

        d2 = (
            d1
            -
            sigma
            * math.sqrt(T)
        )

        return (
            K
            * math.exp(-r * T)
            * normal_cdf(-d2)
            -
            S
            * normal_cdf(-d1)
        )

    except Exception:

        return np.nan


# ============================================================
# IMPLIED VOLATILITY
# ============================================================

def solve_iv(
    option_price,
    S,
    K,
    T,
    r,
    option_type
):

    if not np.isfinite(option_price):

        return np.nan


    if not np.isfinite(S):

        return np.nan


    if (
        S <= 0
        or K <= 0
        or T <= 0
    ):

        return np.nan


    intrinsic = max(
        0.0,
        (
            S - K
            if option_type == "CE"
            else K - S
        )
    )


    if option_price <= intrinsic:

        return np.nan


    low = 0.0001

    high = 5.0


    try:

        if option_type == "CE":

            f_low = (
                bs_call_price(
                    S,
                    K,
                    T,
                    r,
                    low
                )
                -
                option_price
            )

            f_high = (
                bs_call_price(
                    S,
                    K,
                    T,
                    r,
                    high
                )
                -
                option_price
            )

        else:

            f_low = (
                bs_put_price(
                    S,
                    K,
                    T,
                    r,
                    low
                )
                -
                option_price
            )

            f_high = (
                bs_put_price(
                    S,
                    K,
                    T,
                    r,
                    high
                )
                -
                option_price
            )


        if (
            not np.isfinite(f_low)
            or
            not np.isfinite(f_high)
        ):

            return np.nan


        if f_low * f_high > 0:

            return np.nan


        for _ in range(100):

            mid = (
                low
                +
                high
            ) / 2.0


            if option_type == "CE":

                price = bs_call_price(
                    S,
                    K,
                    T,
                    r,
                    mid
                )

            else:

                price = bs_put_price(
                    S,
                    K,
                    T,
                    r,
                    mid
                )


            f_mid = (
                price
                -
                option_price
            )


            if abs(f_mid) < 1e-7:

                return mid


            if f_low * f_mid <= 0:

                high = mid

                f_high = f_mid

            else:

                low = mid

                f_low = f_mid


        return (
            low
            +
            high
        ) / 2.0


    except Exception:

        return np.nan


# ============================================================
# GREEKS
# ============================================================

def calculate_greeks(
    S,
    K,
    T,
    r,
    iv,
    option_type
):

    result = {

        "Delta": np.nan,

        "Gamma": np.nan,

        "Theta": np.nan,

        "Vega": np.nan

    }


    if not all(
        np.isfinite(x)
        for x in [
            S,
            K,
            T,
            r,
            iv
        ]
    ):

        return result


    if (
        S <= 0
        or K <= 0
        or T <= 0
        or iv <= 0
    ):

        return result


    try:

        sqrt_T = math.sqrt(T)


        d1 = (
            math.log(S / K)
            +
            (
                r
                +
                0.5 * iv * iv
            )
            * T
        ) / (
            iv
            * sqrt_T
        )


        d2 = (
            d1
            -
            iv
            * sqrt_T
        )


        gamma = (
            normal_pdf(d1)
            /
            (
                S
                * iv
                * sqrt_T
            )
        )


        vega = (
            S
            * normal_pdf(d1)
            * sqrt_T
            / 100.0
        )


        if option_type == "CE":

            delta = normal_cdf(d1)

            theta = (

                -(
                    S
                    * normal_pdf(d1)
                    * iv
                )
                /
                (
                    2
                    * sqrt_T
                )

                -

                r
                * K
                * math.exp(-r * T)
                * normal_cdf(d2)

            ) / DAYS_PER_YEAR


        else:

            delta = (
                normal_cdf(d1)
                - 1.0
            )


            theta = (

                -(
                    S
                    * normal_pdf(d1)
                    * iv
                )
                /
                (
                    2
                    * sqrt_T
                )

                +

                r
                * K
                * math.exp(-r * T)
                * normal_cdf(-d2)

            ) / DAYS_PER_YEAR


        result = {

            "Delta":
                delta,

            "Gamma":
                gamma,

            "Theta":
                theta,

            "Vega":
                vega

        }


    except Exception:

        pass


    return result


# ============================================================
# HISTORICAL CANDLES
# ============================================================

def get_historical_candles(
    instrument_key,
    from_date,
    to_date
):

    url = (
        API_V3
        +
        "/historical-candle/"
        +
        requests.utils.quote(
            instrument_key,
            safe=""
        )
        +
        "/"
        +
        INTERVAL_UNIT
        +
        "/"
        +
        INTERVAL
        +
        "/"
        +
        to_date
        +
        "/"
        +
        from_date
    )


    data = api_get(
        url
    )


    if not data:

        return []


    try:

        return (
            data
            .get(
                "data",
                {}
            )
            .get(
                "candles",
                []
            )
        )

    except Exception:

        return []


# ============================================================
# HISTORICAL OI
#
# IMPORTANT:
#
# This is the correct API for historical strike-wise OI.
#
# /v2/market/oi
#
# It returns:
#
# call_oi
# put_oi
# strike_price
#
# ============================================================

def get_historical_oi(
    date_string
):

    url = (
        API_V2
        +
        "/market/oi"
    )


    params = {

        "instrument_key":
            UNDERLYING_KEY,

        "expiry":
            EXPIRY_DATE,

        "date":
            date_string

    }


    data = api_get(

        url,

        params=params

    )


    if not data:

        return {}


    try:

        outer_data = data.get(
            "data"
        )


        if not isinstance(
            outer_data,
            dict
        ):

            return {}


        oi_list = outer_data.get(
            "call_put_oi_data_list",
            []
        )


        if not isinstance(
            oi_list,
            list
        ):

            return {}


        result = {}


        for item in oi_list:

            if not isinstance(
                item,
                dict
            ):

                continue


            strike = safe_float(
                item.get(
                    "strike_price"
                )
            )


            call_oi = safe_float(
                item.get(
                    "call_oi"
                )
            )


            put_oi = safe_float(
                item.get(
                    "put_oi"
                )
            )


            if not np.isfinite(
                strike
            ):

                continue


            result[
                float(strike)
            ] = {

                "CE_OI_RAW":
                    call_oi,

                "PE_OI_RAW":
                    put_oi

            }


        return result


    except Exception as error:

        print(
            "Historical OI parse error:",
            error
        )

        return {}


# ============================================================
# GET ACTIVE CONTRACTS
# ============================================================

print("=" * 80)
print("GETTING ACTIVE NIFTY OPTION CONTRACTS")
print("=" * 80)
print()

print(
    "Underlying:",
    UNDERLYING_KEY
)

print(
    "Expiry:",
    EXPIRY_DATE
)

print()


contracts_response = api_get(

    API_V2 + "/option/contract",

    params={

        "instrument_key":
            UNDERLYING_KEY,

        "expiry_date":
            EXPIRY_DATE

    }

)


if contracts_response is None:

    print()
    print(
        "FAILED TO GET OPTION CONTRACTS."
    )

    raise SystemExit


contracts = contracts_response.get(
    "data",
    []
)


print(
    "Contracts received:",
    len(contracts)
)


if not contracts:

    print()
    print(
        "NO OPTION CONTRACTS."
    )

    raise SystemExit


# ============================================================
# PARSE CONTRACTS
# ============================================================

contract_rows = []


for contract in contracts:

    try:

        instrument_key = contract.get(
            "instrument_key"
        )

        strike = contract.get(
            "strike_price"
        )

        option_type = contract.get(
            "instrument_type"
        )

        trading_symbol = contract.get(
            "trading_symbol"
        )

        expiry = contract.get(
            "expiry"
        )

        weekly = contract.get(
            "weekly"
        )

        lot_size = contract.get(
            "lot_size"
        )


        if (
            not instrument_key
            or strike is None
            or option_type is None
        ):

            continue


        option_type = str(
            option_type
        ).upper()


        if option_type not in (
            "CE",
            "PE"
        ):

            continue


        contract_rows.append({

            "instrument_key":
                instrument_key,

            "strike":
                float(strike),

            "option_type":
                option_type,

            "trading_symbol":
                trading_symbol,

            "expiry":
                expiry,

            "weekly":
                weekly,

            "lot_size":
                lot_size

        })


    except Exception:

        continue


contracts_df = pd.DataFrame(
    contract_rows
)


if contracts_df.empty:

    print(
        "NO VALID CE/PE CONTRACTS."
    )

    raise SystemExit


contracts_df = (
    contracts_df
    .drop_duplicates(
        subset=[
            "instrument_key"
        ]
    )
)


print()

print(
    "CE contracts:",
    len(
        contracts_df[
            contracts_df[
                "option_type"
            ] == "CE"
        ]
    )
)

print(
    "PE contracts:",
    len(
        contracts_df[
            contracts_df[
                "option_type"
            ] == "PE"
        ]
    )
)


# ============================================================
# DATE RANGE
# ============================================================

today = datetime.now(
    IST
).date()


start_date_obj = datetime.strptime(
    START_DATE,
    "%Y-%m-%d"
).date()


end_date_obj = datetime.strptime(
    END_DATE,
    "%Y-%m-%d"
).date()


effective_end_date = min(
    end_date_obj,
    today
)


print()
print("=" * 80)
print("DATE RANGE")
print("=" * 80)
print()

print(
    "Today:",
    today
)

print(
    "Start:",
    start_date_obj
)

print(
    "End:",
    effective_end_date
)

print()


if effective_end_date < start_date_obj:

    print(
        "INVALID DATE RANGE."
    )

    raise SystemExit


# ============================================================
# DOWNLOAD SPOT DATA
# ============================================================

print("=" * 80)
print("DOWNLOADING NIFTY 5-MINUTE DATA")
print("=" * 80)
print()


spot_records = []


chunk_start = start_date_obj

chunk_number = 0


while chunk_start <= effective_end_date:

    chunk_end = min(

        chunk_start
        +
        timedelta(days=27),

        effective_end_date

    )


    chunk_number += 1


    print(
        "Spot chunk:",
        chunk_number,
        chunk_start,
        "->",
        chunk_end
    )


    candles = get_historical_candles(

        UNDERLYING_KEY,

        chunk_start.strftime(
            "%Y-%m-%d"
        ),

        chunk_end.strftime(
            "%Y-%m-%d"
        )

    )


    print(
        "Candles:",
        len(candles)
    )


    for candle in candles:

        try:

            timestamp = pd.to_datetime(
                candle[0],
                utc=True
            ).tz_convert(
                IST
            )


            spot_records.append({

                "timestamp":
                    timestamp,

                "spot_open":
                    safe_float(
                        candle[1]
                    ),

                "spot_high":
                    safe_float(
                        candle[2]
                    ),

                "spot_low":
                    safe_float(
                        candle[3]
                    ),

                "spot_close":
                    safe_float(
                        candle[4]
                    ),

                "spot_volume":
                    safe_float(
                        candle[5]
                    )
                    if len(candle) > 5
                    else np.nan,

                "spot_oi":
                    safe_float(
                        candle[6]
                    )
                    if len(candle) > 6
                    else np.nan

            })


        except Exception:

            continue


    chunk_start = (
        chunk_end
        +
        timedelta(days=1)
    )


    time.sleep(
        0.20
    )


print()

print(
    "Total spot candles:",
    len(spot_records)
)


if not spot_records:

    print(
        "NO SPOT DATA."
    )

    raise SystemExit


spot_df = pd.DataFrame(
    spot_records
)


spot_df = (
    spot_df
    .sort_values(
        "timestamp"
    )
    .drop_duplicates(
        "timestamp"
    )
    .reset_index(
        drop=True
    )
)


# ============================================================
# LATEST SPOT
# ============================================================

latest_spot = (
    spot_df[
        "spot_close"
    ]
    .dropna()
    .iloc[-1]
)


print()

print(
    "Latest NIFTY:",
    latest_spot
)


# ============================================================
# ATM
# ============================================================

ATM_STRIKE = (
    round(
        latest_spot
        /
        STRIKE_STEP
    )
    *
    STRIKE_STEP
)


print(
    "ATM Strike:",
    ATM_STRIKE
)


# ============================================================
# SELECT STRIKES
# ============================================================

lower_strike = (
    ATM_STRIKE
    -
    STRIKES_EACH_SIDE
    *
    STRIKE_STEP
)


upper_strike = (
    ATM_STRIKE
    +
    STRIKES_EACH_SIDE
    *
    STRIKE_STEP
)


selected_strikes = sorted(

    [

        strike

        for strike
        in contracts_df[
            "strike"
        ].unique()

        if (
            lower_strike
            <= strike
            <= upper_strike
        )

    ]

)


selected_contracts = (
    contracts_df[
        contracts_df[
            "strike"
        ].isin(
            selected_strikes
        )
    ]
    .copy()
)


print()
print("=" * 80)
print("SELECTED STRIKES")
print("=" * 80)
print()

print(
    "Lower:",
    lower_strike
)

print(
    "ATM:",
    ATM_STRIKE
)

print(
    "Upper:",
    upper_strike
)

print()

print(
    selected_strikes
)

print()

print(
    "Selected contracts:",
    len(selected_contracts)
)


if selected_contracts.empty:

    print(
        "NO CONTRACTS SELECTED."
    )

    raise SystemExit


# ============================================================
# DOWNLOAD OPTION CANDLES
# ============================================================

print()
print("=" * 80)
print("DOWNLOADING OPTION 5-MINUTE DATA")
print("=" * 80)
print()


option_records = []


total_contracts = len(
    selected_contracts
)


for counter, (_, contract) in enumerate(

    selected_contracts.iterrows(),

    1

):

    instrument_key = contract[
        "instrument_key"
    ]

    strike = contract[
        "strike"
    ]

    option_type = contract[
        "option_type"
    ]

    trading_symbol = contract[
        "trading_symbol"
    ]


    print()

    print(
        f"[{counter}/{total_contracts}]",
        trading_symbol
    )

    print(
        "Key:",
        instrument_key
    )


    chunk_start = start_date_obj

    chunk_number = 0


    while chunk_start <= effective_end_date:

        chunk_end = min(

            chunk_start
            +
            timedelta(days=27),

            effective_end_date

        )


        chunk_number += 1


        print(
            "Chunk:",
            chunk_number,
            chunk_start,
            "->",
            chunk_end
        )


        candles = get_historical_candles(

            instrument_key,

            chunk_start.strftime(
                "%Y-%m-%d"
            ),

            chunk_end.strftime(
                "%Y-%m-%d"
            )

        )


        print(
            "Candles:",
            len(candles)
        )


        for candle in candles:

            try:

                timestamp = pd.to_datetime(
                    candle[0],
                    utc=True
                ).tz_convert(
                    IST
                )


                option_records.append({

                    "timestamp":
                        timestamp,

                    "strike":
                        strike,

                    "option_type":
                        option_type,

                    "instrument_key":
                        instrument_key,

                    "trading_symbol":
                        trading_symbol,

                    "open":
                        safe_float(
                            candle[1]
                        ),

                    "high":
                        safe_float(
                            candle[2]
                        ),

                    "low":
                        safe_float(
                            candle[3]
                        ),

                    "close":
                        safe_float(
                            candle[4]
                        ),

                    "volume_raw":
                        safe_float(
                            candle[5]
                        )
                        if len(candle) > 5
                        else np.nan,

                    "candle_oi_raw":
                        safe_float(
                            candle[6]
                        )
                        if len(candle) > 6
                        else np.nan

                })


            except Exception:

                continue


        chunk_start = (
            chunk_end
            +
            timedelta(days=1)
        )


        time.sleep(
            0.20
        )


print()

print(
    "Total raw option candles:",
    len(option_records)
)


if not option_records:

    print(
        "NO OPTION DATA."
    )

    raise SystemExit


# ============================================================
# OPTION DATAFRAME
# ============================================================

options_df = pd.DataFrame(
    option_records
)


options_df = (
    options_df
    .sort_values(
        [
            "timestamp",
            "strike",
            "option_type"
        ]
    )
    .drop_duplicates(
        subset=[
            "timestamp",
            "instrument_key"
        ]
    )
    .reset_index(
        drop=True
    )
)


# ============================================================
# DATE COLUMN
# ============================================================

options_df[
    "trade_date"
] = (

    options_df[
        "timestamp"
    ]
    .dt
    .date

)


# ============================================================
# HISTORICAL OI
#
# Download one OI snapshot for every trading date.
# ============================================================

print()
print("=" * 80)
print("DOWNLOADING HISTORICAL OI")
print("=" * 80)
print()


oi_records = []


current_date = start_date_obj


while current_date <= effective_end_date:

    date_string = current_date.strftime(
        "%Y-%m-%d"
    )


    print(
        "OI date:",
        date_string
    )


    oi_data = get_historical_oi(
        date_string
    )


    if oi_data:

        for strike, values in oi_data.items():

            oi_records.append({

                "trade_date":
                    current_date,

                "strike":
                    float(strike),

                "CE_OI_RAW":
                    values.get(
                        "CE_OI_RAW",
                        np.nan
                    ),

                "PE_OI_RAW":
                    values.get(
                        "PE_OI_RAW",
                        np.nan
                    )

            })


        print(
            "   OI strikes:",
            len(oi_data)
        )


    else:

        print(
            "   OI not returned"
        )


    current_date = (
        current_date
        +
        timedelta(days=1)
    )


    time.sleep(
        0.20
    )


print()

print(
    "Historical OI rows:",
    len(oi_records)
)


# ============================================================
# OI DATAFRAME
# ============================================================

if oi_records:

    oi_df = pd.DataFrame(
        oi_records
    )


    oi_df = (
        oi_df
        .drop_duplicates(
            subset=[
                "trade_date",
                "strike"
            ]
        )
        .reset_index(
            drop=True
        )
    )

else:

    oi_df = pd.DataFrame(

        columns=[
            "trade_date",
            "strike",
            "CE_OI_RAW",
            "PE_OI_RAW"
        ]

    )


# ============================================================
# DAILY OI + DAILY CHANGE OI
#
# This MUST be calculated on the daily OI table.
# Do not calculate change after expanding the daily snapshot
# over 5-minute candles.
# ============================================================

if not oi_df.empty:

    oi_df = oi_df.sort_values(
        [
            "strike",
            "trade_date"
        ]
    ).reset_index(
        drop=True
    )

    oi_df[
        "CE_OI"
    ] = (
        oi_df[
            "CE_OI_RAW"
        ]
        /
        LOT_SIZE
    )

    oi_df[
        "PE_OI"
    ] = (
        oi_df[
            "PE_OI_RAW"
        ]
        /
        LOT_SIZE
    )

    oi_df[
        "CE_Change_OI"
    ] = (
        oi_df
        .groupby(
            "strike"
        )[
            "CE_OI"
        ]
        .diff()
    )

    oi_df[
        "PE_Change_OI"
    ] = (
        oi_df
        .groupby(
            "strike"
        )[
            "PE_OI"
        ]
        .diff()
    )


# ============================================================
# MERGE HISTORICAL OI INTO OPTION DATA
# ============================================================

print()
print(
    "Merging historical OI..."
)


options_df = options_df.merge(

    oi_df,

    on=[
        "trade_date",
        "strike"
    ],

    how="left"

)


# ============================================================
# FALLBACK
#
# If historical /market/oi is unavailable for a row,
# use candle OI only as fallback.
#
# But the primary OI remains /market/oi.
# ============================================================

options_df[
    "CE_OI_RAW"
] = np.where(

    options_df[
        "option_type"
    ] == "CE",

    options_df[
        "CE_OI_RAW"
    ],

    np.nan

)


options_df[
    "PE_OI_RAW"
] = np.where(

    options_df[
        "option_type"
    ] == "PE",

    options_df[
        "PE_OI_RAW"
    ],

    np.nan

)


# ============================================================
# IMPORTANT:
#
# For each option row:
#
# CE_OI_RAW is populated only for CE
# PE_OI_RAW is populated only for PE
#
# Now convert raw OI into NSE lot-style quantity.
# ============================================================

options_df[
    "OI_RAW"
] = np.where(

    options_df[
        "option_type"
    ] == "CE",

    options_df[
        "CE_OI_RAW"
    ],

    options_df[
        "PE_OI_RAW"
    ]

)


options_df[
    "OI"
] = (

    options_df[
        "OI_RAW"
    ]
    /
    LOT_SIZE

)


# ============================================================
# MERGE SPOT
# ============================================================

print()
print("=" * 80)
print("MERGING NIFTY + OPTION DATA")
print("=" * 80)
print()


options_df = options_df.sort_values(
    "timestamp"
)

spot_df = spot_df.sort_values(
    "timestamp"
)


merged = pd.merge_asof(

    options_df,

    spot_df,

    on="timestamp",

    direction="backward"

)


# ============================================================
# IV / GREEKS
# ============================================================

print()
print(
    "Calculating IV + Greeks..."
)


expiry_datetime = datetime.strptime(
    EXPIRY_DATE,
    "%Y-%m-%d"
).replace(
    hour=15,
    minute=30,
    tzinfo=IST
)


iv_values = []

delta_values = []

gamma_values = []

theta_values = []

vega_values = []

dte_values = []


for _, row in merged.iterrows():

    S = safe_float(
        row.get(
            "spot_close"
        )
    )

    K = safe_float(
        row.get(
            "strike"
        )
    )

    option_price = safe_float(
        row.get(
            "close"
        )
    )

    option_type = row[
        "option_type"
    ]

    timestamp = row[
        "timestamp"
    ]


    seconds = (
        expiry_datetime
        -
        timestamp
    ).total_seconds()


    dte = max(

        seconds
        /
        (
            24
            * 60
            * 60
        ),

        0.0

    )


    dte_values.append(
        dte
    )


    T = max(

        seconds
        /
        (
            DAYS_PER_YEAR
            * 24
            * 60
            * 60
        ),

        1.0 / DAYS_PER_YEAR

    )


    iv = solve_iv(

        option_price,

        S,

        K,

        T,

        RISK_FREE_RATE,

        option_type

    )


    if np.isfinite(iv):

        iv_percent = (
            iv * 100.0
        )

    else:

        iv_percent = np.nan


    iv_values.append(
        iv_percent
    )


    greeks = calculate_greeks(

        S,

        K,

        T,

        RISK_FREE_RATE,

        iv,

        option_type

    )


    delta_values.append(
        greeks["Delta"]
    )

    gamma_values.append(
        greeks["Gamma"]
    )

    theta_values.append(
        greeks["Theta"]
    )

    vega_values.append(
        greeks["Vega"]
    )


merged[
    "DTE"
] = dte_values


merged[
    "IV"
] = iv_values


merged[
    "Delta"
] = delta_values


merged[
    "Gamma"
] = gamma_values


merged[
    "Theta"
] = theta_values


merged[
    "Vega"
] = vega_values


# ============================================================
# DAILY VOLUME
#
# Sum all 5-minute candle volume for each option/date.
#
# This is the important part for comparing with daily
# NSE TOT_TRADED_QTY.
# ============================================================

print()
print(
    "Calculating daily traded quantity..."
)


merged[
    "Daily_Volume_RAW"
] = (

    merged
    .groupby(
        [
            "trade_date",
            "instrument_key"
        ]
    )[
        "volume_raw"
    ]
    .transform(
        "sum"
    )

)


merged[
    "Daily_Volume"
] = (

    merged[
        "Daily_Volume_RAW"
    ]
    /
    LOT_SIZE

)


# ============================================================
# CE / PE
# ============================================================

ce = merged[
    merged[
        "option_type"
    ] == "CE"
].copy()


pe = merged[
    merged[
        "option_type"
    ] == "PE"
].copy()


# ============================================================
# CE OI
#
# CE_OI_RAW = daily raw Upstox historical OI
# CE_OI     = CE_OI_RAW / LOT_SIZE
# CE_Change_OI = DAILY OI CHANGE, not 5-minute diff
#
# Historical OI from /market/oi is a daily snapshot.
# The same daily OI is therefore repeated on every 5-minute
# candle for that trading date.
# Change OI must be calculated from the daily OI table BEFORE
# it is merged into the 5-minute option candles.
# ============================================================

ce[
    "CE_OI"
] = (

    ce[
        "CE_OI_RAW"
    ]
    /
    LOT_SIZE

)


ce[
    "CE_Volume"
] = (

    ce[
        "volume_raw"
    ]
    /
    LOT_SIZE

)


ce[
    "CE_Daily_TOT_TRADED_QTY"
] = ce[
    "Daily_Volume_RAW"
]


ce[
    "CE_Daily_Volume"
] = ce[
    "Daily_Volume"
]


# ============================================================
# PE OI
# ============================================================

pe[
    "PE_OI"
] = (

    pe[
        "PE_OI_RAW"
    ]
    /
    LOT_SIZE

)


pe[
    "PE_Volume"
] = (

    pe[
        "volume_raw"
    ]
    /
    LOT_SIZE

)


pe[
    "PE_Daily_TOT_TRADED_QTY"
] = pe[
    "Daily_Volume_RAW"
]


pe[
    "PE_Daily_Volume"
] = pe[
    "Daily_Volume"
]


# ============================================================
# RENAME CE
# ============================================================

ce = ce.rename(

    columns={

        "open":
            "CE_Open",

        "high":
            "CE_High",

        "low":
            "CE_Low",

        "close":
            "CE_Close",

        "volume_raw":
            "CE_Volume_RAW",

        "CE_OI_RAW":
            "CE_OI_RAW",

        "CE_IV":
            "CE_IV",

        "IV":
            "CE_IV",

        "Delta":
            "CE_Delta",

        "Gamma":
            "CE_Gamma",

        "Theta":
            "CE_Theta",

        "Vega":
            "CE_Vega",

        "DTE":
            "CE_DTE"

    }

)


# ============================================================
# RENAME PE
# ============================================================

pe = pe.rename(

    columns={

        "open":
            "PE_Open",

        "high":
            "PE_High",

        "low":
            "PE_Low",

        "close":
            "PE_Close",

        "volume_raw":
            "PE_Volume_RAW",

        "IV":
            "PE_IV",

        "Delta":
            "PE_Delta",

        "Gamma":
            "PE_Gamma",

        "Theta":
            "PE_Theta",

        "Vega":
            "PE_Vega",

        "DTE":
            "PE_DTE"

    }

)


# ============================================================
# CE COLUMNS
# ============================================================

ce_columns = [

    "timestamp",

    "trade_date",

    "strike",

    "spot_open",

    "spot_high",

    "spot_low",

    "spot_close",

    "spot_volume",

    "spot_oi",

    "CE_Open",

    "CE_High",

    "CE_Low",

    "CE_Close",

    "CE_Volume_RAW",

    "CE_Volume",

    "CE_Daily_TOT_TRADED_QTY",

    "CE_Daily_Volume",

    "CE_OI_RAW",

    "CE_OI",

    "CE_Change_OI",

    "CE_IV",

    "CE_Delta",

    "CE_Gamma",

    "CE_Theta",

    "CE_Vega",

    "CE_DTE",

    "trading_symbol",

    "instrument_key"

]


ce_columns = [

    column

    for column
    in ce_columns

    if column in ce.columns

]


ce = ce[
    ce_columns
]


# ============================================================
# PE COLUMNS
# ============================================================

pe_columns = [

    "timestamp",

    "trade_date",

    "strike",

    "PE_Open",

    "PE_High",

    "PE_Low",

    "PE_Close",

    "PE_Volume_RAW",

    "PE_Volume",

    "PE_Daily_TOT_TRADED_QTY",

    "PE_Daily_Volume",

    "PE_OI_RAW",

    "PE_OI",

    "PE_Change_OI",

    "PE_IV",

    "PE_Delta",

    "PE_Gamma",

    "PE_Theta",

    "PE_Vega",

    "PE_DTE",

    "trading_symbol",

    "instrument_key"

]


pe_columns = [

    column

    for column
    in pe_columns

    if column in pe.columns

]


pe = pe[
    pe_columns
]


# ============================================================
# RENAME CONTRACT INFO
# ============================================================

ce = ce.rename(

    columns={

        "trading_symbol":
            "CE_Symbol",

        "instrument_key":
            "CE_Instrument_Key"

    }

)


pe = pe.rename(

    columns={

        "trading_symbol":
            "PE_Symbol",

        "instrument_key":
            "PE_Instrument_Key"

    }

)


# ============================================================
# MERGE CE + PE
# ============================================================

chain = pd.merge(

    ce,

    pe,

    on=[
        "timestamp",
        "strike"
    ],

    how="outer"

)


# ============================================================
# SORT
# ============================================================

chain = chain.sort_values(

    [
        "timestamp",
        "strike"
    ]

).reset_index(
    drop=True
)


# ============================================================
# CHANGE OI
#
# IMPORTANT FIX
#
# CE_OI and PE_OI are DAILY OI snapshots.
# They are repeated across all 5-minute candles for a date.
# Therefore a 5-minute diff would produce 0 on almost every row.
#
# We already calculated the DAILY change in oi_df above.
# Just carry that daily change into the final chain.
#
# CE_Change_OI = today's CE_OI - previous trading day's CE_OI
# PE_Change_OI = today's PE_OI - previous trading day's PE_OI
# ============================================================

chain = chain.sort_values(

    [
        "timestamp",
        "strike"
    ]

).reset_index(
    drop=True
)


# Safety: if the daily change columns were not created,
# create them as NaN instead of generating incorrect 5-minute diffs.

if "CE_Change_OI" not in chain.columns:

    chain[
        "CE_Change_OI"
    ] = np.nan


if "PE_Change_OI" not in chain.columns:

    chain[
        "PE_Change_OI"
    ] = np.nan


# ============================================================
# PCR
#
# PUT OI / CALL OI
# ============================================================

chain[
    "PCR_OI"
] = np.where(

    chain[
        "CE_OI"
    ].notna()
    &
    (
        chain[
            "CE_OI"
        ].abs() > 0
    ),

    chain[
        "PE_OI"
    ]
    /
    chain[
        "CE_OI"
    ],

    np.nan

)


# ============================================================
# PCR CHANGE OI
# ============================================================

chain[
    "PCR_Change_OI"
] = np.where(

    chain[
        "CE_Change_OI"
    ].notna()
    &
    (
        chain[
            "CE_Change_OI"
        ].abs() > 0
    ),

    chain[
        "PE_Change_OI"
    ]
    /
    chain[
        "CE_Change_OI"
    ],

    np.nan

)


# ============================================================
# TOTAL OI
# ============================================================

chain[
    "Total_OI"
] = (

    chain[
        "CE_OI"
    ].fillna(0)

    +

    chain[
        "PE_OI"
    ].fillna(0)

)


# ============================================================
# TOTAL RAW OI
# ============================================================

chain[
    "Total_OI_RAW"
] = (

    chain[
        "CE_OI_RAW"
    ].fillna(0)

    +

    chain[
        "PE_OI_RAW"
    ].fillna(0)

)


# ============================================================
# OI DIFFERENCE
# ============================================================

chain[
    "OI_Difference"
] = (

    chain[
        "PE_OI"
    ]

    -

    chain[
        "CE_OI"
    ]

)


# ============================================================
# IV DIFFERENCE
# ============================================================

chain[
    "IV_Difference"
] = (

    chain[
        "PE_IV"
    ]

    -

    chain[
        "CE_IV"
    ]

)


# ============================================================
# ATM DISTANCE
# ============================================================

chain[
    "ATM_Distance"
] = (

    chain[
        "strike"
    ]

    -

    chain[
        "spot_close"
    ]

)


# ============================================================
# MONEYNESS
# ============================================================

def get_moneyness(
    strike,
    spot
):

    if (
        not np.isfinite(strike)
        or
        not np.isfinite(spot)
    ):

        return ""


    if abs(
        strike - spot
    ) <= 25:

        return "ATM"


    if strike < spot:

        return "ITM"


    return "OTM"


chain[
    "Moneyness"
] = [

    get_moneyness(
        row["strike"],
        row["spot_close"]
    )

    for _, row
    in chain.iterrows()

]


# ============================================================
# FINAL SORT
# ============================================================

chain = chain.sort_values(

    [
        "timestamp",
        "strike"
    ]

).reset_index(
    drop=True
)


# ============================================================
# FINAL COLUMN ORDER
# ============================================================

final_columns = [

    "timestamp",

    "trade_date",

    # SPOT
    "spot_open",
    "spot_high",
    "spot_low",
    "spot_close",
    "spot_volume",
    "spot_oi",

    # CALL
    "CE_Open",
    "CE_High",
    "CE_Low",
    "CE_Close",

    "CE_Volume_RAW",
    "CE_Volume",

    "CE_Daily_TOT_TRADED_QTY",
    "CE_Daily_Volume",

    "CE_OI_RAW",
    "CE_OI",
    "CE_Change_OI",

    "CE_IV",
    "CE_Delta",
    "CE_Gamma",
    "CE_Theta",
    "CE_Vega",
    "CE_DTE",

    # STRIKE
    "strike",
    "Moneyness",
    "ATM_Distance",

    # OI
    "PCR_OI",
    "PCR_Change_OI",

    "Total_OI_RAW",
    "Total_OI",

    "OI_Difference",

    "IV_Difference",

    # PUT
    "PE_Open",
    "PE_High",
    "PE_Low",
    "PE_Close",

    "PE_Volume_RAW",
    "PE_Volume",

    "PE_Daily_TOT_TRADED_QTY",
    "PE_Daily_Volume",

    "PE_OI_RAW",
    "PE_OI",
    "PE_Change_OI",

    "PE_IV",
    "PE_Delta",
    "PE_Gamma",
    "PE_Theta",
    "PE_Vega",
    "PE_DTE"

]


final_columns = [

    column

    for column
    in final_columns

    if column in chain.columns

]


chain = chain[
    final_columns
]


# ============================================================
# CLEAN INF
# ============================================================

chain = chain.replace(

    [
        np.inf,
        -np.inf
    ],

    np.nan

)


spot_excel = spot_df.copy()

spot_excel = spot_excel.replace(

    [
        np.inf,
        -np.inf
    ],

    np.nan

)


contracts_excel = (
    selected_contracts
    .copy()
)


oi_excel = oi_df.copy()


# ============================================================
# EXCEL DATETIME
# ============================================================

def make_excel_datetime_naive(
    df,
    column
):

    if column not in df.columns:

        return


    dt = pd.to_datetime(
        df[column],
        errors="coerce"
    )


    try:

        if dt.dt.tz is not None:

            dt = (
                dt
                .dt
                .tz_localize(
                    None
                )
            )

    except Exception:

        pass


    df[column] = dt


make_excel_datetime_naive(
    chain,
    "timestamp"
)

make_excel_datetime_naive(
    spot_excel,
    "timestamp"
)


chain = chain.reset_index(
    drop=True
)

spot_excel = spot_excel.reset_index(
    drop=True
)


# ============================================================
# SUMMARY
# ============================================================

summary = pd.DataFrame({

    "Parameter": [

        "Underlying",

        "Underlying Key",

        "Expiry",

        "Requested Start",

        "Requested End",

        "Effective End",

        "Interval",

        "Lot Size",

        "Latest Historical Spot",

        "ATM Strike",

        "Lower Strike",

        "Upper Strike",

        "Selected Strikes",

        "CE Contracts",

        "PE Contracts",

        "Raw Option Candles",

        "Historical OI Rows",

        "Final Chain Rows",

        "Risk Free Rate"

    ],


    "Value": [

        UNDERLYING_NAME,

        UNDERLYING_KEY,

        EXPIRY_DATE,

        START_DATE,

        END_DATE,

        effective_end_date.strftime(
            "%Y-%m-%d"
        ),

        "5 minutes",

        LOT_SIZE,

        latest_spot,

        ATM_STRIKE,

        lower_strike,

        upper_strike,

        len(
            selected_strikes
        ),

        len(
            selected_contracts[
                selected_contracts[
                    "option_type"
                ] == "CE"
            ]
        ),

        len(
            selected_contracts[
                selected_contracts[
                    "option_type"
                ] == "PE"
            ]
        ),

        len(
            option_records
        ),

        len(
            oi_df
        ),

        len(
            chain
        ),

        RISK_FREE_RATE

    ]

})


# ============================================================
# ALL-STRIKE OPTION CHAIN + EXPIRIES
# ============================================================

print()
print("=" * 80)
print("DOWNLOADING ALL-STRIKE OPTION CHAIN")
print("=" * 80)
print()

all_strike_chain = build_all_strike_chain(EXPIRY_DATE)

print(
    "All-strike rows:",
    len(all_strike_chain)
)

print()
print("=" * 80)
print("DOWNLOADING PAST / COMING EXPIRIES")
print("=" * 80)
print()

expiry_df, active_expiries, past_expiries = build_expiry_dataframe()

print(
    "Active/coming expiries:",
    len(active_expiries)
)
print(
    "Past expiries:",
    len(past_expiries)
)

coming_expiry_df = expiry_df[
    expiry_df["Status"].isin(["TODAY", "COMING"])
].copy() if not expiry_df.empty else pd.DataFrame()

past_expiry_df = expiry_df[
    expiry_df["Status"] == "PAST"
].copy() if not expiry_df.empty else pd.DataFrame()


# ============================================================
# SQLITE DATABASE
# ============================================================

print()
print("=" * 80)
print("SAVING DATA TO SQLITE DATABASE")
print("=" * 80)
print()

try:
    db = sqlite3.connect(DB_FILE)

    chain.to_sql(
        "optionchain_5min",
        db,
        if_exists="replace",
        index=False
    )

    spot_excel.to_sql(
        "nifty_spot_5min",
        db,
        if_exists="replace",
        index=False
    )

    contracts_excel.to_sql(
        "contracts_selected",
        db,
        if_exists="replace",
        index=False
    )

    oi_excel.to_sql(
        "historical_oi",
        db,
        if_exists="replace",
        index=False
    )

    all_strike_chain.to_sql(
        "optionchain_all_strikes",
        db,
        if_exists="replace",
        index=False
    )

    expiry_df.to_sql(
        "expiry_list",
        db,
        if_exists="replace",
        index=False
    )

    coming_expiry_df.to_sql(
        "coming_expiries",
        db,
        if_exists="replace",
        index=False
    )

    past_expiry_df.to_sql(
        "past_expiries",
        db,
        if_exists="replace",
        index=False
    )

    db.commit()
    db.close()

    print("Database saved:")
    print(DB_FILE)

except Exception as error:
    print()
    print("DATABASE ERROR")
    print(error)


# ============================================================
# WRITE EXCEL
# ============================================================

print()
print("=" * 80)
print("WRITING EXCEL")
print("=" * 80)
print()


try:

    with pd.ExcelWriter(

        OUTPUT_FILE,

        engine="openpyxl"

    ) as writer:

        chain.to_excel(

            writer,

            sheet_name="OptionChain_5Min",

            index=False

        )


        spot_excel.to_excel(

            writer,

            sheet_name="NIFTY_Spot_5Min",

            index=False

        )


        contracts_excel.to_excel(

            writer,

            sheet_name="Contracts",

            index=False

        )


        oi_excel.to_excel(

            writer,

            sheet_name="Historical_OI",

            index=False

        )


        all_strike_chain.to_excel(

            writer,

            sheet_name="OptionChain_AllStrikes",

            index=False

        )


        expiry_df.to_excel(

            writer,

            sheet_name="Expiry_List",

            index=False

        )


        coming_expiry_df.to_excel(

            writer,

            sheet_name="Coming_Expiries",

            index=False

        )


        past_expiry_df.to_excel(

            writer,

            sheet_name="Past_Expiries",

            index=False

        )


        summary.to_excel(

            writer,

            sheet_name="Summary",

            index=False

        )


except Exception as error:

    print()
    print("=" * 80)
    print("EXCEL ERROR")
    print("=" * 80)
    print()

    print(
        error
    )

    traceback.print_exc()

    raise SystemExit


# ============================================================
# FORMAT EXCEL
# ============================================================

print(
    "Formatting Excel..."
)


try:

    from openpyxl import load_workbook

    from openpyxl.styles import Font

    from openpyxl.styles import Alignment

    from openpyxl.utils import get_column_letter


    wb = load_workbook(
        OUTPUT_FILE
    )


    for ws in wb.worksheets:

        ws.freeze_panes = "A2"


        if ws.max_row > 1:

            ws.auto_filter.ref = (
                ws.dimensions
            )


        for cell in ws[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )


        for col_idx in range(

            1,

            ws.max_column + 1

        ):

            letter = get_column_letter(
                col_idx
            )


            max_length = 0


            for cell in ws[letter]:

                try:

                    length = len(
                        str(
                            cell.value
                        )
                    )


                    if length > max_length:

                        max_length = length

                except Exception:

                    pass


            ws.column_dimensions[
                letter
            ].width = min(

                max(
                    max_length + 2,
                    10
                ),

                32

            )


    # ========================================================
    # MAIN SHEET
    # ========================================================

    ws = wb[
        "OptionChain_5Min"
    ]


    header_map = {

        cell.value:
            cell.column

        for cell in ws[1]

    }


    # ========================================================
    # DATETIME
    # ========================================================

    for column_name in [

        "timestamp"

    ]:

        if column_name not in header_map:

            continue


        col = get_column_letter(

            header_map[
                column_name
            ]

        )


        for cell in ws[col][1:]:

            cell.number_format = (
                "yyyy-mm-dd hh:mm:ss"
            )


    # ========================================================
    # DECIMAL
    # ========================================================

    decimal_columns = [

        "spot_open",
        "spot_high",
        "spot_low",
        "spot_close",

        "CE_Open",
        "CE_High",
        "CE_Low",
        "CE_Close",

        "PE_Open",
        "PE_High",
        "PE_Low",
        "PE_Close",

        "CE_IV",
        "PE_IV",

        "CE_Delta",
        "PE_Delta",

        "CE_Gamma",
        "PE_Gamma",

        "CE_Theta",
        "PE_Theta",

        "CE_Vega",
        "PE_Vega",

        "PCR_OI",
        "PCR_Change_OI",

        "IV_Difference",

        "ATM_Distance"

    ]


    for column_name in decimal_columns:

        if column_name not in header_map:

            continue


        col = get_column_letter(

            header_map[
                column_name
            ]

        )


        for cell in ws[col][1:]:

            cell.number_format = (
                "0.0000"
            )


    # ========================================================
    # INTEGER
    # ========================================================

    integer_columns = [

        "spot_volume",
        "spot_oi",

        "CE_Volume_RAW",
        "CE_Volume",

        "CE_Daily_TOT_TRADED_QTY",
        "CE_Daily_Volume",

        "CE_OI_RAW",
        "CE_OI",
        "CE_Change_OI",

        "PE_Volume_RAW",
        "PE_Volume",

        "PE_Daily_TOT_TRADED_QTY",
        "PE_Daily_Volume",

        "PE_OI_RAW",
        "PE_OI",
        "PE_Change_OI",

        "Total_OI_RAW",
        "Total_OI",

        "OI_Difference"

    ]


    for column_name in integer_columns:

        if column_name not in header_map:

            continue


        col = get_column_letter(

            header_map[
                column_name
            ]

        )


        for cell in ws[col][1:]:

            cell.number_format = (
                "#,##0"
            )


    # ========================================================
    # STRIKE
    # ========================================================

    if "strike" in header_map:

        col = get_column_letter(

            header_map[
                "strike"
            ]

        )


        for cell in ws[col][1:]:

            cell.number_format = (
                "0"
            )


    # ========================================================
    # HISTORICAL OI SHEET
    # ========================================================

    if "Historical_OI" in wb.sheetnames:

        ws_oi = wb[
            "Historical_OI"
        ]


        oi_header_map = {

            cell.value:
                cell.column

            for cell in ws_oi[1]

        }


        for column_name in [

            "CE_OI_RAW",
            "PE_OI_RAW"

        ]:

            if column_name not in oi_header_map:

                continue


            col = get_column_letter(

                oi_header_map[
                    column_name
                ]

            )


            for cell in ws_oi[col][1:]:

                cell.number_format = (
                    "#,##0"
                )


    # ========================================================
    # ALL-STRIKE / EXPIRY SHEETS
    # ========================================================

    for sheet_name in [
        "OptionChain_AllStrikes",
        "Expiry_List",
        "Coming_Expiries",
        "Past_Expiries"
    ]:

        if sheet_name not in wb.sheetnames:
            continue

        ws_extra = wb[sheet_name]
        extra_header = {
            cell.value: cell.column
            for cell in ws_extra[1]
        }

        for name in [
            "strike",
            "underlying_spot",
            "CE_LTP",
            "CE_Close",
            "PE_LTP",
            "PE_Close",
            "CE_IV",
            "PE_IV",
            "CE_Delta",
            "PE_Delta",
            "CE_Gamma",
            "PE_Gamma",
            "CE_Theta",
            "PE_Theta",
            "CE_Vega",
            "PE_Vega",
            "PCR_OI",
            "PCR_Change_OI"
        ]:
            if name not in extra_header:
                continue
            col = get_column_letter(extra_header[name])
            for cell in ws_extra[col][1:]:
                cell.number_format = "0.0000"

        for name in [
            "CE_Volume_RAW",
            "CE_Volume",
            "CE_OI_RAW",
            "CE_OI",
            "CE_Previous_OI_RAW",
            "CE_Previous_OI",
            "CE_Change_OI",
            "CE_Bid_Qty",
            "CE_Ask_Qty",
            "PE_Volume_RAW",
            "PE_Volume",
            "PE_OI_RAW",
            "PE_OI",
            "PE_Previous_OI_RAW",
            "PE_Previous_OI",
            "PE_Change_OI",
            "PE_Bid_Qty",
            "PE_Ask_Qty",
            "Total_OI",
            "OI_Difference"
        ]:
            if name not in extra_header:
                continue
            col = get_column_letter(extra_header[name])
            for cell in ws_extra[col][1:]:
                cell.number_format = "#,##0"


    wb.save(
        OUTPUT_FILE
    )


except Exception as error:

    print()
    print(
        "Excel formatting warning:"
    )

    print(
        error
    )


# ============================================================
# TEST 24200
# ============================================================

print()
print("=" * 80)
print("OI / VOLUME CHECK")
print("=" * 80)
print()


test_rows = chain[
    chain[
        "strike"
    ] == 24200
]


if not test_rows.empty:

    test_row = test_rows.iloc[-1]


    print(
        "Strike:",
        test_row["strike"]
    )

    print()

    print(
        "CE OI RAW:",
        test_row.get(
            "CE_OI_RAW"
        )
    )

    print(
        "CE OI / 65:",
        test_row.get(
            "CE_OI"
        )
    )

    print(
        "CE Change OI:",
        test_row.get(
            "CE_Change_OI"
        )
    )

    print()

    print(
        "PE OI RAW:",
        test_row.get(
            "PE_OI_RAW"
        )
    )

    print(
        "PE OI / 65:",
        test_row.get(
            "PE_OI"
        )
    )

    print(
        "PE Change OI:",
        test_row.get(
            "PE_Change_OI"
        )
    )

    print()

    print(
        "CE 5-min Volume RAW:",
        test_row.get(
            "CE_Volume_RAW"
        )
    )

    print(
        "CE 5-min Volume / 65:",
        test_row.get(
            "CE_Volume"
        )
    )

    print()

    print(
        "CE Daily TOT TRADED QTY:",
        test_row.get(
            "CE_Daily_TOT_TRADED_QTY"
        )
    )

    print(
        "CE Daily Volume / 65:",
        test_row.get(
            "CE_Daily_Volume"
        )
    )

    print()

    print(
        "PE Daily TOT TRADED QTY:",
        test_row.get(
            "PE_Daily_TOT_TRADED_QTY"
        )
    )

    print(
        "PE Daily Volume / 65:",
        test_row.get(
            "PE_Daily_Volume"
        )
    )

else:

    print(
        "24200 strike not found."
    )


# ============================================================
# FINAL
# ============================================================

print()
print("=" * 80)
print("SUCCESS")
print("=" * 80)
print()

print(
    "Excel file:"
)

print(
    OUTPUT_FILE
)

print()
print("Database:")
print(DB_FILE)

print()

print(
    "Final rows:",
    len(chain)
)

print(
    "Final columns:",
    len(chain.columns)
)

print()

print(
    "Sheets:"
)

print(
    "1. OptionChain_5Min"
)

print(
    "2. NIFTY_Spot_5Min"
)

print(
    "3. Contracts"
)

print(
    "4. Historical_OI"
)

print(
    "5. OptionChain_AllStrikes"
)

print(
    "6. Expiry_List"
)

print(
    "7. Coming_Expiries"
)

print(
    "8. Past_Expiries"
)

print(
    "9. Summary"
)

print()

print("=" * 80)
print("DONE")
print("=" * 80)

