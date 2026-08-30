# ============================================================
# UPSTOX NIFTY 5-MIN HISTORICAL OPTION CHAIN
# ============================================================
#
# Python 3.14
#
# INSTALL:
#
# pip install requests pandas numpy openpyxl
#
# TOKEN:
#
# C:\Users\soiku\Desktop\access_token.txt
#
# OUTPUT:
#
# C:\Users\soiku\Desktop\Upstox_NIFTY_5Min_OptionChain.xlsx
#
# ============================================================
#
# IMPORTANT OI LOGIC
#
# CE_OI_RAW       = Upstox historical raw OI
# PE_OI_RAW       = Upstox historical raw OI
#
# CE_OI           = CE_OI_RAW / 65
# PE_OI           = PE_OI_RAW / 65
#
# CE_Change_OI    = daily CE_OI - previous trading day CE_OI
# PE_Change_OI    = daily PE_OI - previous trading day PE_OI
#
# PCR_OI          = PE_OI / CE_OI
#
# IMPORTANT VOLUME LOGIC
#
# CE_Volume_RAW   = Upstox 5-minute candle volume
# PE_Volume_RAW   = Upstox 5-minute candle volume
#
# CE_Daily_Volume_RAW
#                 = sum of 5-minute raw volumes for that
#                   option on that trading date
#
# CE_Daily_Volume
#                 = CE_Daily_Volume_RAW / 65
#
# Example:
#
# 355007705 / 65 = 5461657
#
# ============================================================

import os
import math
import time
import traceback
import sqlite3

from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import requests
import numpy as np
import pandas as pd


# ============================================================
# SETTINGS
# ============================================================

BASE_DIR = r"C:\Users\soiku\Desktop"

TOKEN_FILE = os.path.join(
    BASE_DIR,
    "access_token.txt"
)

OUTPUT_FILE = os.path.join(
    BASE_DIR,
    "Upstox_NIFTY_5Min_OptionChain.xlsx"
)

DB_FILE = os.path.join(
    BASE_DIR,
    "Upstox_NIFTY_OptionChain.db"
)


# ============================================================
# UPSTOX API
# ============================================================

API_V2 = "https://api.upstox.com/v2"

API_V3 = "https://api.upstox.com/v3"


# ============================================================
# NIFTY
# ============================================================

UNDERLYING_KEY = "NSE_INDEX|Nifty 50"

UNDERLYING_NAME = "NIFTY"


# ============================================================
# EXPIRY
# ============================================================

EXPIRY_DATE = "2026-09-01"

# ============================================================
# HISTORICAL RANGE
# ============================================================

START_DATE = "2026-08-13"

END_DATE = "2026-08-28"

# ============================================================
# 5 MINUTE
# ============================================================

INTERVAL_UNIT = "minutes"

INTERVAL = "5"


# ============================================================
# LOT SIZE
# ============================================================

LOT_SIZE = 65


# ============================================================
# STRIKES
#
# ATM +/- 10 strikes
# ============================================================

STRIKES_EACH_SIDE = 10

STRIKE_STEP = 50


# ============================================================
# NSE-STYLE OPTION CHAIN -- EXTRA DATES
#
# The NSE-style sheet for TODAY is always built (full columns:
# OI, Change OI, Volume, IV, LTP, Change, Bid/Ask).
#
# Type any additional past dates here (format "YYYY-MM-DD") to
# get one extra NSE-style sheet per date, e.g.:
#
#   EXTRA_OI_DATES = ["2026-08-28", "2026-08-27"]
# ============================================================

EXTRA_OI_DATES = ["2026-08-27", "2026-08-28", "2026-08-26"]


# ============================================================
# IV
# ============================================================

RISK_FREE_RATE = 0.10

DAYS_PER_YEAR = 365.0


# ============================================================
# TIMEZONE
# ============================================================

IST = ZoneInfo(
    "Asia/Kolkata"
)


# ============================================================
# DISPLAY
# ============================================================

pd.set_option(
    "display.max_columns",
    None
)

pd.set_option(
    "display.width",
    300
)

pd.set_option(
    "display.max_rows",
    100
)


# ============================================================
# HEADER
# ============================================================

print()
print("=" * 80)
print("UPSTOX NIFTY 5-MIN HISTORICAL OPTION CHAIN")
print("=" * 80)
print()

print(
    "Underlying :",
    UNDERLYING_KEY
)

print(
    "Expiry     :",
    EXPIRY_DATE
)

print(
    "History    :",
    START_DATE,
    "->",
    END_DATE
)

print(
    "Interval   :",
    "5 minutes"
)

print(
    "Lot Size   :",
    LOT_SIZE
)

print()

print(
    "Output     :",
    OUTPUT_FILE
)

print(
    "Database   :",
    DB_FILE
)

print()


# ============================================================
# LOAD TOKEN
# ============================================================

print("=" * 80)
print("LOADING ACCESS TOKEN")
print("=" * 80)
print()


if not os.path.exists(TOKEN_FILE):

    print("ERROR:")
    print(
        "access_token.txt was not found."
    )

    print()
    print(
        "Create:"
    )

    print(
        TOKEN_FILE
    )

    raise SystemExit


with open(
    TOKEN_FILE,
    "r",
    encoding="utf-8"
) as f:

    ACCESS_TOKEN = f.read().strip()


if not ACCESS_TOKEN:

    print("ERROR:")
    print(
        "access_token.txt is empty."
    )

    raise SystemExit


print(
    "Access token loaded."
)

print()


# ============================================================
# HTTP SESSION
# ============================================================

session = requests.Session()

session.headers.update(
    {
        "Authorization":
            f"Bearer {ACCESS_TOKEN}",

        "Accept":
            "application/json",

        "Content-Type":
            "application/json"
    }
)


# ============================================================
# API GET
# ============================================================

def api_get(
    url,
    params=None,
    timeout=30
):

    try:

        response = session.get(
            url,
            params=params,
            timeout=timeout
        )

    except Exception as error:

        print()
        print(
            "NETWORK ERROR"
        )

        print(
            error
        )

        return None


    if response.status_code != 200:

        print()
        print(
            "UPSTOX API ERROR"
        )

        print(
            "HTTP:",
            response.status_code
        )

        print(
            "URL:",
            response.url
        )

        try:

            print(
                response.json()
            )

        except Exception:

            print(
                response.text[:1000]
            )

        return None


    try:

        return response.json()

    except Exception as error:

        print()
        print(
            "JSON ERROR"
        )

        print(
            error
        )

        return None


# ============================================================
# ALL-STRIKE OPTION CHAIN
# ============================================================
# Upstox provides a dedicated option-chain endpoint containing
# every strike for the requested expiry, with CE/PE market data
# and Greeks.
# ============================================================

def get_option_chain(expiry_date):

    url = API_V2 + "/option/chain"

    data = api_get(
        url,
        params={
            "instrument_key": UNDERLYING_KEY,
            "expiry_date": expiry_date
        }
    )

    if not data:
        return []

    try:
        rows = data.get("data", [])
        return rows if isinstance(rows, list) else []
    except Exception:
        return []


def get_iv_reference_datetime():

    # ------------------------------------------------------------
    # WHICH TIMESTAMP THE CURRENT LTP ACTUALLY BELONGS TO.
    #
    # Live LTP (CE_LTP/PE_LTP) is only "live" while the market is
    # open. Outside market hours (evenings, weekends, holidays)
    # Upstox just keeps returning the LAST TRADED price frozen
    # from the last session -- but plain datetime.now(IST) keeps
    # ticking forward. Solving IV with a stale price + a fresh
    # "now" understates time-to-expiry and skews IV (e.g. showing
    # 14.4% instead of the correct ~9.5% once the weekend rolls
    # by). So IV must be solved against the timestamp the price
    # actually corresponds to, not raw wall-clock time.
    #
    # Heuristic (no holiday calendar available here):
    #   - Weekday, 9:15-15:30 IST -> market is live, use now.
    #   - Weekday, before 9:15 IST -> still shows previous
    #     session's close, roll back to the previous weekday
    #     at 15:30.
    #   - Weekday, after 15:30 IST -> today's own close, use
    #     today at 15:30.
    #   - Saturday / Sunday -> roll back to the most recent
    #     Friday at 15:30.
    # ------------------------------------------------------------

    now_ts = datetime.now(IST)

    market_open = now_ts.replace(hour=9, minute=15, second=0, microsecond=0)
    market_close = now_ts.replace(hour=15, minute=30, second=0, microsecond=0)

    if now_ts.weekday() >= 5:
        # Saturday (5) or Sunday (6) -> roll back to Friday close.
        days_back = now_ts.weekday() - 4
        reference = (now_ts - timedelta(days=days_back)).replace(
            hour=15, minute=30, second=0, microsecond=0
        )
        return reference

    if now_ts < market_open:
        # Before today's open -> previous weekday's close.
        days_back = 3 if now_ts.weekday() == 0 else 1
        reference = (now_ts - timedelta(days=days_back)).replace(
            hour=15, minute=30, second=0, microsecond=0
        )
        return reference

    if now_ts > market_close:
        # After today's close -> today's own close.
        return market_close

    # Within market hours -> live.
    return now_ts


def build_all_strike_chain(expiry_date):

    raw_rows = get_option_chain(expiry_date)
    records = []

    for row in raw_rows:

        try:
            strike = safe_float(row.get("strike_price"))
            spot = safe_float(row.get("underlying_spot_price"))
            expiry = row.get("expiry", expiry_date)
            pcr = safe_float(row.get("pcr"))

            if not np.isfinite(strike):
                continue

            call = row.get("call_options") or {}
            put = row.get("put_options") or {}

            cm = call.get("market_data") or {}
            pm = put.get("market_data") or {}
            cg = call.get("option_greeks") or {}
            pg = put.get("option_greeks") or {}

            call_oi_raw = safe_float(cm.get("oi"))
            put_oi_raw = safe_float(pm.get("oi"))
            call_prev_oi_raw = safe_float(cm.get("prev_oi"))
            put_prev_oi_raw = safe_float(pm.get("prev_oi"))

            records.append({
                "expiry": expiry,
                "underlying_spot": spot,
                "strike": strike,
                "Moneyness": get_moneyness(strike, spot) if np.isfinite(spot) else "",

                "CE_Instrument_Key": call.get("instrument_key"),
                "CE_LTP": safe_float(cm.get("ltp")),
                "CE_Close": safe_float(cm.get("close_price")),
                "CE_Volume_RAW": safe_float(cm.get("volume")),
                "CE_Volume": safe_float(cm.get("volume")) / LOT_SIZE,
                "CE_OI_RAW": call_oi_raw,
                "CE_OI": call_oi_raw / LOT_SIZE if np.isfinite(call_oi_raw) else np.nan,
                "CE_Previous_OI_RAW": call_prev_oi_raw,
                "CE_Previous_OI": call_prev_oi_raw / LOT_SIZE if np.isfinite(call_prev_oi_raw) else np.nan,
                "CE_Change_OI": (call_oi_raw - call_prev_oi_raw) / LOT_SIZE if np.isfinite(call_oi_raw) and np.isfinite(call_prev_oi_raw) else np.nan,
                "CE_Bid": safe_float(cm.get("bid_price")),
                "CE_Bid_Qty": safe_float(cm.get("bid_qty")),
                "CE_Ask": safe_float(cm.get("ask_price")),
                "CE_Ask_Qty": safe_float(cm.get("ask_qty")),
                "CE_IV": safe_float(cg.get("iv")),
                "CE_Delta": safe_float(cg.get("delta")),
                "CE_Gamma": safe_float(cg.get("gamma")),
                "CE_Theta": safe_float(cg.get("theta")),
                "CE_Vega": safe_float(cg.get("vega")),
                "CE_POP": safe_float(cg.get("pop")),

                "PCR_OI": pcr,

                "PE_Instrument_Key": put.get("instrument_key"),
                "PE_LTP": safe_float(pm.get("ltp")),
                "PE_Close": safe_float(pm.get("close_price")),
                "PE_Volume_RAW": safe_float(pm.get("volume")),
                "PE_Volume": safe_float(pm.get("volume")) / LOT_SIZE,
                "PE_OI_RAW": put_oi_raw,
                "PE_OI": put_oi_raw / LOT_SIZE if np.isfinite(put_oi_raw) else np.nan,
                "PE_Previous_OI_RAW": put_prev_oi_raw,
                "PE_Previous_OI": put_prev_oi_raw / LOT_SIZE if np.isfinite(put_prev_oi_raw) else np.nan,
                "PE_Change_OI": (put_oi_raw - put_prev_oi_raw) / LOT_SIZE if np.isfinite(put_oi_raw) and np.isfinite(put_prev_oi_raw) else np.nan,
                "PE_Bid": safe_float(pm.get("bid_price")),
                "PE_Bid_Qty": safe_float(pm.get("bid_qty")),
                "PE_Ask": safe_float(pm.get("ask_price")),
                "PE_Ask_Qty": safe_float(pm.get("ask_qty")),
                "PE_IV": safe_float(pg.get("iv")),
                "PE_Delta": safe_float(pg.get("delta")),
                "PE_Gamma": safe_float(pg.get("gamma")),
                "PE_Theta": safe_float(pg.get("theta")),
                "PE_Vega": safe_float(pg.get("vega")),
                "PE_POP": safe_float(pg.get("pop"))
            })

        except Exception:
            continue

    df = pd.DataFrame(records)

    if not df.empty:
        df = df.sort_values("strike").reset_index(drop=True)

        # ------------------------------------------------------------
        # RECOMPUTE CE_IV / PE_IV WITH THE SAME BLACK-SCHOLES SOLVER
        # USED ON OptionChain_5Min (solve_iv), instead of trusting
        # Upstox's own "iv" field.
        #
        # OptionChain_5Min already solves IV itself and that has been
        # matching NSE's displayed IV accurately, while Upstox's raw
        # greeks IV can sit slightly off (e.g. 9.75 vs NSE/solve_iv's
        # 9.5256 at the same strike). Using the exact same solver
        # here keeps every sheet consistent with the one that was
        # already accurate.
        # ------------------------------------------------------------

        expiry_datetime = datetime.strptime(
            expiry_date, "%Y-%m-%d"
        ).replace(hour=15, minute=30, tzinfo=IST)

        now_ts = get_iv_reference_datetime()

        seconds_to_expiry = (
            expiry_datetime - now_ts
        ).total_seconds()

        T = max(
            seconds_to_expiry / (DAYS_PER_YEAR * 24 * 60 * 60),
            1.0 / DAYS_PER_YEAR
        )

        recomputed_ce_iv = []
        recomputed_pe_iv = []

        for _, row in df.iterrows():

            S = row.get("underlying_spot")
            K = row.get("strike")

            ce_price = row.get("CE_LTP")
            pe_price = row.get("PE_LTP")

            ce_iv = solve_iv(ce_price, S, K, T, RISK_FREE_RATE, "CE")
            pe_iv = solve_iv(pe_price, S, K, T, RISK_FREE_RATE, "PE")

            recomputed_ce_iv.append(
                ce_iv * 100.0 if np.isfinite(ce_iv) else np.nan
            )
            recomputed_pe_iv.append(
                pe_iv * 100.0 if np.isfinite(pe_iv) else np.nan
            )

        df["CE_IV"] = recomputed_ce_iv
        df["PE_IV"] = recomputed_pe_iv

        df["Total_OI"] = df["CE_OI"].fillna(0) + df["PE_OI"].fillna(0)
        df["OI_Difference"] = df["PE_OI"] - df["CE_OI"]
        df["PCR_Change_OI"] = np.where(
            df["CE_Change_OI"].notna() & (df["CE_Change_OI"].abs() > 0),
            df["PE_Change_OI"] / df["CE_Change_OI"],
            np.nan
        )

    return df


# ============================================================
# EXPIRY LIST
# ============================================================

def get_active_expiries():

    data = api_get(
        API_V2 + "/option/contract",
        params={"instrument_key": UNDERLYING_KEY}
    )

    expiries = set()

    if data:
        for item in data.get("data", []):
            expiry = item.get("expiry")
            if expiry:
                expiries.add(str(expiry))

    return sorted(expiries)


def get_past_expiries():

    data = api_get(
        API_V2 + "/expired-instruments/expiries",
        params={"instrument_key": UNDERLYING_KEY}
    )

    if not data:
        return []

    values = data.get("data", [])
    return sorted({str(x) for x in values if x})


def build_expiry_dataframe():

    today_date = datetime.now(IST).date()

    active = get_active_expiries()
    past = get_past_expiries()

    rows = []
    all_dates = sorted(set(active + past))

    for expiry in all_dates:
        try:
            expiry_date = datetime.strptime(expiry, "%Y-%m-%d").date()
        except Exception:
            continue

        if expiry_date < today_date:
            status = "PAST"
            source = "Upstox expired-instruments"
        elif expiry_date == today_date:
            status = "TODAY"
            source = "Upstox active contracts"
        else:
            status = "COMING"
            source = "Upstox active contracts"

        rows.append({
            "Expiry": expiry,
            "Status": status,
            "Source": source
        })

    df = pd.DataFrame(rows)

    if not df.empty:
        df["Expiry"] = pd.to_datetime(df["Expiry"]).dt.date.astype(str)
        df = df.sort_values("Expiry").reset_index(drop=True)

    return df, active, past


# ============================================================
# SAFE FLOAT
# ============================================================

def safe_float(value):

    try:

        if value is None:

            return np.nan

        return float(value)

    except Exception:

        return np.nan


# ============================================================
# NORMAL CDF
# ============================================================

def normal_cdf(x):

    return (
        0.5
        *
        (
            1.0
            +
            math.erf(
                x / math.sqrt(2.0)
            )
        )
    )


# ============================================================
# NORMAL PDF
# ============================================================

def normal_pdf(x):

    return (
        math.exp(
            -0.5 * x * x
        )
        /
        math.sqrt(
            2.0 * math.pi
        )
    )


# ============================================================
# BLACK-SCHOLES CALL
# ============================================================

def bs_call_price(
    S,
    K,
    T,
    r,
    sigma
):

    if (
        S <= 0
        or K <= 0
        or T <= 0
        or sigma <= 0
    ):

        return np.nan


    try:

        d1 = (
            math.log(S / K)
            +
            (
                r
                +
                0.5 * sigma * sigma
            )
            * T
        ) / (
            sigma
            * math.sqrt(T)
        )

        d2 = (
            d1
            -
            sigma
            * math.sqrt(T)
        )

        return (
            S * normal_cdf(d1)
            -
            K
            * math.exp(-r * T)
            * normal_cdf(d2)
        )

    except Exception:

        return np.nan


# ============================================================
# BLACK-SCHOLES PUT
# ============================================================

def bs_put_price(
    S,
    K,
    T,
    r,
    sigma
):

    if (
        S <= 0
        or K <= 0
        or T <= 0
        or sigma <= 0
    ):

        return np.nan


    try:

        d1 = (
            math.log(S / K)
            +
            (
                r
                +
                0.5 * sigma * sigma
            )
            * T
        ) / (
            sigma
            * math.sqrt(T)
        )

        d2 = (
            d1
            -
            sigma
            * math.sqrt(T)
        )

        return (
            K
            * math.exp(-r * T)
            * normal_cdf(-d2)
            -
            S
            * normal_cdf(-d1)
        )

    except Exception:

        return np.nan


# ============================================================
# IMPLIED VOLATILITY
# ============================================================

def solve_iv(
    option_price,
    S,
    K,
    T,
    r,
    option_type
):

    if not np.isfinite(option_price):

        return np.nan


    if not np.isfinite(S):

        return np.nan


    if (
        S <= 0
        or K <= 0
        or T <= 0
    ):

        return np.nan


    intrinsic = max(
        0.0,
        (
            S - K
            if option_type == "CE"
            else K - S
        )
    )


    if option_price <= intrinsic:

        return np.nan


    low = 0.0001

    high = 5.0


    try:

        if option_type == "CE":

            f_low = (
                bs_call_price(
                    S,
                    K,
                    T,
                    r,
                    low
                )
                -
                option_price
            )

            f_high = (
                bs_call_price(
                    S,
                    K,
                    T,
                    r,
                    high
                )
                -
                option_price
            )

        else:

            f_low = (
                bs_put_price(
                    S,
                    K,
                    T,
                    r,
                    low
                )
                -
                option_price
            )

            f_high = (
                bs_put_price(
                    S,
                    K,
                    T,
                    r,
                    high
                )
                -
                option_price
            )


        if (
            not np.isfinite(f_low)
            or
            not np.isfinite(f_high)
        ):

            return np.nan


        if f_low * f_high > 0:

            return np.nan


        for _ in range(100):

            mid = (
                low
                +
                high
            ) / 2.0


            if option_type == "CE":

                price = bs_call_price(
                    S,
                    K,
                    T,
                    r,
                    mid
                )

            else:

                price = bs_put_price(
                    S,
                    K,
                    T,
                    r,
                    mid
                )


            f_mid = (
                price
                -
                option_price
            )


            if abs(f_mid) < 1e-7:

                return mid


            if f_low * f_mid <= 0:

                high = mid

                f_high = f_mid

            else:

                low = mid

                f_low = f_mid


        return (
            low
            +
            high
        ) / 2.0


    except Exception:

        return np.nan


# ============================================================
# GREEKS
# ============================================================

def calculate_greeks(
    S,
    K,
    T,
    r,
    iv,
    option_type
):

    result = {

        "Delta": np.nan,

        "Gamma": np.nan,

        "Theta": np.nan,

        "Vega": np.nan

    }


    if not all(
        np.isfinite(x)
        for x in [
            S,
            K,
            T,
            r,
            iv
        ]
    ):

        return result


    if (
        S <= 0
        or K <= 0
        or T <= 0
        or iv <= 0
    ):

        return result


    try:

        sqrt_T = math.sqrt(T)


        d1 = (
            math.log(S / K)
            +
            (
                r
                +
                0.5 * iv * iv
            )
            * T
        ) / (
            iv
            * sqrt_T
        )


        d2 = (
            d1
            -
            iv
            * sqrt_T
        )


        gamma = (
            normal_pdf(d1)
            /
            (
                S
                * iv
                * sqrt_T
            )
        )


        vega = (
            S
            * normal_pdf(d1)
            * sqrt_T
            / 100.0
        )


        if option_type == "CE":

            delta = normal_cdf(d1)

            theta = (

                -(
                    S
                    * normal_pdf(d1)
                    * iv
                )
                /
                (
                    2
                    * sqrt_T
                )

                -

                r
                * K
                * math.exp(-r * T)
                * normal_cdf(d2)

            ) / DAYS_PER_YEAR


        else:

            delta = (
                normal_cdf(d1)
                - 1.0
            )


            theta = (

                -(
                    S
                    * normal_pdf(d1)
                    * iv
                )
                /
                (
                    2
                    * sqrt_T
                )

                +

                r
                * K
                * math.exp(-r * T)
                * normal_cdf(-d2)

            ) / DAYS_PER_YEAR


        result = {

            "Delta":
                delta,

            "Gamma":
                gamma,

            "Theta":
                theta,

            "Vega":
                vega

        }


    except Exception:

        pass


    return result


# ============================================================
# HISTORICAL CANDLES
# ============================================================

def get_historical_candles(
    instrument_key,
    from_date,
    to_date,
    is_spot=False
):

    # ------------------------------------------------------------
    # The underlying index (NIFTY 50 spot) never "expires", so it
    # always uses the live v3 historical-candle endpoint.
    #
    # An OPTION instrument, once its expiry has passed, moves to
    # Upstox's separate expired-instruments endpoint -- the live
    # endpoint returns nothing for it.
    # ------------------------------------------------------------

    if is_spot:

        url = (
            API_V3
            +
            "/historical-candle/"
            +
            requests.utils.quote(
                instrument_key,
                safe=""
            )
            +
            "/"
            +
            INTERVAL_UNIT
            +
            "/"
            +
            INTERVAL
            +
            "/"
            +
            to_date
            +
            "/"
            +
            from_date
        )

    elif IS_PAST_EXPIRY:

        url = (
            API_V2
            +
            "/expired-instruments/historical-candle/"
            +
            requests.utils.quote(
                instrument_key,
                safe=""
            )
            +
            "/"
            +
            INTERVAL
            +
            "minute/"
            +
            to_date
            +
            "/"
            +
            from_date
        )

    else:

        url = (
            API_V3
            +
            "/historical-candle/"
            +
            requests.utils.quote(
                instrument_key,
                safe=""
            )
            +
            "/"
            +
            INTERVAL_UNIT
            +
            "/"
            +
            INTERVAL
            +
            "/"
            +
            to_date
            +
            "/"
            +
            from_date
        )


    data = api_get(
        url
    )


    if not data:

        return []


    try:

        return (
            data
            .get(
                "data",
                {}
            )
            .get(
                "candles",
                []
            )
        )

    except Exception:

        return []


# ============================================================
# HISTORICAL OI
#
# IMPORTANT:
#
# This is the correct API for historical strike-wise OI.
#
# /v2/market/oi
#
# It returns:
#
# call_oi
# put_oi
# strike_price
#
# ============================================================

def get_historical_oi(
    date_string
):

    url = (
        API_V2
        +
        "/market/oi"
    )


    params = {

        "instrument_key":
            UNDERLYING_KEY,

        "expiry":
            EXPIRY_DATE,

        "date":
            date_string

    }


    data = api_get(

        url,

        params=params

    )


    if not data:

        return {}


    try:

        outer_data = data.get(
            "data"
        )


        if not isinstance(
            outer_data,
            dict
        ):

            return {}


        oi_list = outer_data.get(
            "call_put_oi_data_list",
            []
        )


        if not isinstance(
            oi_list,
            list
        ):

            return {}


        result = {}


        for item in oi_list:

            if not isinstance(
                item,
                dict
            ):

                continue


            strike = safe_float(
                item.get(
                    "strike_price"
                )
            )


            call_oi = safe_float(
                item.get(
                    "call_oi"
                )
            )


            put_oi = safe_float(
                item.get(
                    "put_oi"
                )
            )


            if not np.isfinite(
                strike
            ):

                continue


            result[
                float(strike)
            ] = {

                "CE_OI_RAW":
                    call_oi,

                "PE_OI_RAW":
                    put_oi

            }


        return result


    except Exception as error:

        print(
            "Historical OI parse error:",
            error
        )

        return {}


# ============================================================
# EXPIRY TYPE DETECTION (PAST vs ACTIVE/FUTURE)
#
# Upstox's LIVE endpoints (/option/contract, /historical-candle,
# /option/chain) only cover contracts that are still active or
# upcoming. Once an expiry has passed, those contracts move to
# the separate "expired-instruments" endpoints. Using the live
# endpoints for a past EXPIRY_DATE returns nothing, which is why
# the script previously failed whenever a past expiry was typed
# in. This flag routes contract/candle downloads to the correct
# endpoint set, and switches the all-strike chain source (no live
# snapshot exists for an expired contract).
# ============================================================

_today_for_expiry_check = datetime.now(IST).date()

try:
    _expiry_date_obj = datetime.strptime(
        EXPIRY_DATE, "%Y-%m-%d"
    ).date()
    IS_PAST_EXPIRY = _expiry_date_obj < _today_for_expiry_check
except Exception:
    print("Invalid EXPIRY_DATE format. Use YYYY-MM-DD")
    raise SystemExit

print()
print(
    "Expiry type:",
    "PAST" if IS_PAST_EXPIRY else "ACTIVE/FUTURE"
)


# ============================================================
# GET ACTIVE CONTRACTS
# ============================================================

print("=" * 80)
print("GETTING ACTIVE NIFTY OPTION CONTRACTS")
print("=" * 80)
print()

print(
    "Underlying:",
    UNDERLYING_KEY
)

print(
    "Expiry:",
    EXPIRY_DATE
)

print()


if IS_PAST_EXPIRY:

    contracts_response = api_get(

        API_V2 + "/expired-instruments/option/contract",

        params={

            "instrument_key":
                UNDERLYING_KEY,

            "expiry_date":
                EXPIRY_DATE

        }

    )

else:

    contracts_response = api_get(

        API_V2 + "/option/contract",

        params={

            "instrument_key":
                UNDERLYING_KEY,

            "expiry_date":
                EXPIRY_DATE

        }

    )


if contracts_response is None:

    print()
    print(
        "FAILED TO GET OPTION CONTRACTS."
    )

    raise SystemExit


contracts = contracts_response.get(
    "data",
    []
)


print(
    "Contracts received:",
    len(contracts)
)


if not contracts:

    print()
    print(
        "NO OPTION CONTRACTS."
    )

    raise SystemExit


# ============================================================
# PARSE CONTRACTS
# ============================================================

contract_rows = []


for contract in contracts:

    try:

        instrument_key = contract.get(
            "instrument_key"
        )

        strike = contract.get(
            "strike_price"
        )

        option_type = contract.get(
            "instrument_type"
        )

        trading_symbol = contract.get(
            "trading_symbol"
        )

        expiry = contract.get(
            "expiry"
        )

        weekly = contract.get(
            "weekly"
        )

        lot_size = contract.get(
            "lot_size"
        )


        if (
            not instrument_key
            or strike is None
            or option_type is None
        ):

            continue


        option_type = str(
            option_type
        ).upper()


        if option_type not in (
            "CE",
            "PE"
        ):

            continue


        contract_rows.append({

            "instrument_key":
                instrument_key,

            "strike":
                float(strike),

            "option_type":
                option_type,

            "trading_symbol":
                trading_symbol,

            "expiry":
                expiry,

            "weekly":
                weekly,

            "lot_size":
                lot_size

        })


    except Exception:

        continue


contracts_df = pd.DataFrame(
    contract_rows
)


if contracts_df.empty:

    print(
        "NO VALID CE/PE CONTRACTS."
    )

    raise SystemExit


contracts_df = (
    contracts_df
    .drop_duplicates(
        subset=[
            "instrument_key"
        ]
    )
)


print()

print(
    "CE contracts:",
    len(
        contracts_df[
            contracts_df[
                "option_type"
            ] == "CE"
        ]
    )
)

print(
    "PE contracts:",
    len(
        contracts_df[
            contracts_df[
                "option_type"
            ] == "PE"
        ]
    )
)


# ============================================================
# DATE RANGE
# ============================================================

today = datetime.now(
    IST
).date()


start_date_obj = datetime.strptime(
    START_DATE,
    "%Y-%m-%d"
).date()


end_date_obj = datetime.strptime(
    END_DATE,
    "%Y-%m-%d"
).date()


effective_end_date = min(
    end_date_obj,
    today
)


print()
print("=" * 80)
print("DATE RANGE")
print("=" * 80)
print()

print(
    "Today:",
    today
)

print(
    "Start:",
    start_date_obj
)

print(
    "End:",
    effective_end_date
)

print()


if effective_end_date < start_date_obj:

    print(
        "INVALID DATE RANGE."
    )

    raise SystemExit


# ============================================================
# DOWNLOAD SPOT DATA
# ============================================================

print("=" * 80)
print("DOWNLOADING NIFTY 5-MINUTE DATA")
print("=" * 80)
print()


spot_records = []


chunk_start = start_date_obj

chunk_number = 0


while chunk_start <= effective_end_date:

    chunk_end = min(

        chunk_start
        +
        timedelta(days=27),

        effective_end_date

    )


    chunk_number += 1


    print(
        "Spot chunk:",
        chunk_number,
        chunk_start,
        "->",
        chunk_end
    )


    candles = get_historical_candles(

        UNDERLYING_KEY,

        chunk_start.strftime(
            "%Y-%m-%d"
        ),

        chunk_end.strftime(
            "%Y-%m-%d"
        ),

        is_spot=True

    )


    print(
        "Candles:",
        len(candles)
    )


    for candle in candles:

        try:

            timestamp = pd.to_datetime(
                candle[0],
                utc=True
            ).tz_convert(
                IST
            )


            spot_records.append({

                "timestamp":
                    timestamp,

                "spot_open":
                    safe_float(
                        candle[1]
                    ),

                "spot_high":
                    safe_float(
                        candle[2]
                    ),

                "spot_low":
                    safe_float(
                        candle[3]
                    ),

                "spot_close":
                    safe_float(
                        candle[4]
                    ),

                "spot_volume":
                    safe_float(
                        candle[5]
                    )
                    if len(candle) > 5
                    else np.nan,

                "spot_oi":
                    safe_float(
                        candle[6]
                    )
                    if len(candle) > 6
                    else np.nan

            })


        except Exception:

            continue


    chunk_start = (
        chunk_end
        +
        timedelta(days=1)
    )


    time.sleep(
        0.20
    )


print()

print(
    "Total spot candles:",
    len(spot_records)
)


if not spot_records:

    print(
        "NO SPOT DATA."
    )

    raise SystemExit


spot_df = pd.DataFrame(
    spot_records
)


spot_df = (
    spot_df
    .sort_values(
        "timestamp"
    )
    .drop_duplicates(
        "timestamp"
    )
    .reset_index(
        drop=True
    )
)


# ============================================================
# LATEST SPOT
# ============================================================

latest_spot = (
    spot_df[
        "spot_close"
    ]
    .dropna()
    .iloc[-1]
)


print()

print(
    "Latest NIFTY:",
    latest_spot
)


# ============================================================
# ATM
# ============================================================

ATM_STRIKE = (
    round(
        latest_spot
        /
        STRIKE_STEP
    )
    *
    STRIKE_STEP
)


print(
    "ATM Strike:",
    ATM_STRIKE
)


# ============================================================
# SELECT STRIKES
# ============================================================

lower_strike = (
    ATM_STRIKE
    -
    STRIKES_EACH_SIDE
    *
    STRIKE_STEP
)


upper_strike = (
    ATM_STRIKE
    +
    STRIKES_EACH_SIDE
    *
    STRIKE_STEP
)


selected_strikes = sorted(

    [

        strike

        for strike
        in contracts_df[
            "strike"
        ].unique()

        if (
            lower_strike
            <= strike
            <= upper_strike
        )

    ]

)


selected_contracts = (
    contracts_df[
        contracts_df[
            "strike"
        ].isin(
            selected_strikes
        )
    ]
    .copy()
)


print()
print("=" * 80)
print("SELECTED STRIKES")
print("=" * 80)
print()

print(
    "Lower:",
    lower_strike
)

print(
    "ATM:",
    ATM_STRIKE
)

print(
    "Upper:",
    upper_strike
)

print()

print(
    selected_strikes
)

print()

print(
    "Selected contracts:",
    len(selected_contracts)
)


if selected_contracts.empty:

    print(
        "NO CONTRACTS SELECTED."
    )

    raise SystemExit


# ============================================================
# DOWNLOAD OPTION CANDLES
# ============================================================

print()
print("=" * 80)
print("DOWNLOADING OPTION 5-MINUTE DATA")
print("=" * 80)
print()


option_records = []


total_contracts = len(
    selected_contracts
)


for counter, (_, contract) in enumerate(

    selected_contracts.iterrows(),

    1

):

    instrument_key = contract[
        "instrument_key"
    ]

    strike = contract[
        "strike"
    ]

    option_type = contract[
        "option_type"
    ]

    trading_symbol = contract[
        "trading_symbol"
    ]


    print()

    print(
        f"[{counter}/{total_contracts}]",
        trading_symbol
    )

    print(
        "Key:",
        instrument_key
    )


    chunk_start = start_date_obj

    chunk_number = 0


    while chunk_start <= effective_end_date:

        chunk_end = min(

            chunk_start
            +
            timedelta(days=27),

            effective_end_date

        )


        chunk_number += 1


        print(
            "Chunk:",
            chunk_number,
            chunk_start,
            "->",
            chunk_end
        )


        candles = get_historical_candles(

            instrument_key,

            chunk_start.strftime(
                "%Y-%m-%d"
            ),

            chunk_end.strftime(
                "%Y-%m-%d"
            ),

            is_spot=False

        )


        print(
            "Candles:",
            len(candles)
        )


        for candle in candles:

            try:

                timestamp = pd.to_datetime(
                    candle[0],
                    utc=True
                ).tz_convert(
                    IST
                )


                option_records.append({

                    "timestamp":
                        timestamp,

                    "strike":
                        strike,

                    "option_type":
                        option_type,

                    "instrument_key":
                        instrument_key,

                    "trading_symbol":
                        trading_symbol,

                    "open":
                        safe_float(
                            candle[1]
                        ),

                    "high":
                        safe_float(
                            candle[2]
                        ),

                    "low":
                        safe_float(
                            candle[3]
                        ),

                    "close":
                        safe_float(
                            candle[4]
                        ),

                    "volume_raw":
                        safe_float(
                            candle[5]
                        )
                        if len(candle) > 5
                        else np.nan,

                    "candle_oi_raw":
                        safe_float(
                            candle[6]
                        )
                        if len(candle) > 6
                        else np.nan

                })


            except Exception:

                continue


        chunk_start = (
            chunk_end
            +
            timedelta(days=1)
        )


        time.sleep(
            0.20
        )


print()

print(
    "Total raw option candles:",
    len(option_records)
)


if not option_records:

    print(
        "NO OPTION DATA."
    )

    raise SystemExit


# ============================================================
# OPTION DATAFRAME
# ============================================================

options_df = pd.DataFrame(
    option_records
)


options_df = (
    options_df
    .sort_values(
        [
            "timestamp",
            "strike",
            "option_type"
        ]
    )
    .drop_duplicates(
        subset=[
            "timestamp",
            "instrument_key"
        ]
    )
    .reset_index(
        drop=True
    )
)


# ============================================================
# DATE COLUMN
# ============================================================

options_df[
    "trade_date"
] = (

    options_df[
        "timestamp"
    ]
    .dt
    .date

)


# ============================================================
# HISTORICAL OI
#
# Download one OI snapshot for every trading date.
# ============================================================

print()
print("=" * 80)
print("DOWNLOADING HISTORICAL OI")
print("=" * 80)
print()


oi_records = []


current_date = start_date_obj


while current_date <= effective_end_date:

    date_string = current_date.strftime(
        "%Y-%m-%d"
    )


    print(
        "OI date:",
        date_string
    )


    oi_data = get_historical_oi(
        date_string
    )


    if oi_data:

        for strike, values in oi_data.items():

            oi_records.append({

                "trade_date":
                    current_date,

                "expiry":
                    EXPIRY_DATE,

                "strike":
                    float(strike),

                "CE_OI_RAW":
                    values.get(
                        "CE_OI_RAW",
                        np.nan
                    ),

                "PE_OI_RAW":
                    values.get(
                        "PE_OI_RAW",
                        np.nan
                    )

            })


        print(
            "   OI strikes:",
            len(oi_data)
        )


    else:

        print(
            "   OI not returned"
        )


    current_date = (
        current_date
        +
        timedelta(days=1)
    )


    time.sleep(
        0.20
    )


print()

print(
    "Historical OI rows:",
    len(oi_records)
)


# ============================================================
# OI DATAFRAME
# ============================================================

if oi_records:

    oi_df = pd.DataFrame(
        oi_records
    )


    oi_df = (
        oi_df
        .drop_duplicates(
            subset=[
                "trade_date",
                "strike"
            ]
        )
        .reset_index(
            drop=True
        )
    )

else:

    oi_df = pd.DataFrame(

        columns=[
            "trade_date",
            "expiry",
            "strike",
            "CE_OI_RAW",
            "PE_OI_RAW"
        ]

    )


# ============================================================
# DROP STALE / DUPLICATE OI SNAPSHOT DATES
#
# The historical OI endpoint is often called for every calendar
# date in range, including weekends and holidays. On a
# non-trading date it frequently just echoes back the OI
# snapshot of the last real trading day instead of returning
# nothing. If that duplicate date is left in place it can get
# picked as the "latest" or "previous" date in the sync step,
# which makes CE_Change_OI / PE_Change_OI come out as 0 and can
# also make the reported Date look wrong (e.g. a weekend date).
#
# Fix: for each date (in order), compare the FULL set of raw OI
# values across all strikes to the immediately preceding date.
# If it is byte-for-byte identical, treat it as a stale repeat
# of the same trading session and drop it.
# ============================================================

if not oi_df.empty:

    _sig_df = oi_df.sort_values(
        ["trade_date", "strike"]
    )

    _signatures = _sig_df.groupby("trade_date").apply(
        lambda g: tuple(g["CE_OI_RAW"].fillna(-1)) + tuple(g["PE_OI_RAW"].fillna(-1))
    ).sort_index()

    _is_stale_repeat = _signatures == _signatures.shift(1)

    _stale_dates = set(
        _signatures.index[_is_stale_repeat.fillna(False)]
    )

    if _stale_dates:

        print()
        print(
            "Skipping stale/non-trading OI snapshot dates (identical to previous session):",
            sorted(str(d) for d in _stale_dates)
        )

        oi_df = oi_df[
            ~oi_df["trade_date"].isin(_stale_dates)
        ].reset_index(drop=True)


# ============================================================
# DAILY OI + DAILY CHANGE OI
#
# This MUST be calculated on the daily OI table.
# Do not calculate change after expanding the daily snapshot
# over 5-minute candles.
# ============================================================

if not oi_df.empty:

    oi_df = oi_df.sort_values(
        [
            "strike",
            "trade_date"
        ]
    ).reset_index(
        drop=True
    )

    oi_df[
        "CE_OI"
    ] = (
        oi_df[
            "CE_OI_RAW"
        ]
        /
        LOT_SIZE
    )

    oi_df[
        "PE_OI"
    ] = (
        oi_df[
            "PE_OI_RAW"
        ]
        /
        LOT_SIZE
    )

    oi_df[
        "CE_Change_OI"
    ] = (
        oi_df
        .groupby(
            "strike"
        )[
            "CE_OI"
        ]
        .diff()
    )

    oi_df[
        "PE_Change_OI"
    ] = (
        oi_df
        .groupby(
            "strike"
        )[
            "PE_OI"
        ]
        .diff()
    )


# ============================================================
# MERGE HISTORICAL OI INTO OPTION DATA
# ============================================================

print()
print(
    "Merging historical OI..."
)


options_df = options_df.merge(

    oi_df,

    on=[
        "trade_date",
        "strike"
    ],

    how="left"

)


# ============================================================
# FALLBACK
#
# If historical /market/oi is unavailable for a row,
# use candle OI only as fallback.
#
# But the primary OI remains /market/oi.
# ============================================================

options_df[
    "CE_OI_RAW"
] = np.where(

    options_df[
        "option_type"
    ] == "CE",

    options_df[
        "CE_OI_RAW"
    ],

    np.nan

)


options_df[
    "PE_OI_RAW"
] = np.where(

    options_df[
        "option_type"
    ] == "PE",

    options_df[
        "PE_OI_RAW"
    ],

    np.nan

)


# ============================================================
# IMPORTANT:
#
# For each option row:
#
# CE_OI_RAW is populated only for CE
# PE_OI_RAW is populated only for PE
#
# Now convert raw OI into NSE lot-style quantity.
# ============================================================

options_df[
    "OI_RAW"
] = np.where(

    options_df[
        "option_type"
    ] == "CE",

    options_df[
        "CE_OI_RAW"
    ],

    options_df[
        "PE_OI_RAW"
    ]

)


options_df[
    "OI"
] = (

    options_df[
        "OI_RAW"
    ]
    /
    LOT_SIZE

)


# ============================================================
# MERGE SPOT
# ============================================================

print()
print("=" * 80)
print("MERGING NIFTY + OPTION DATA")
print("=" * 80)
print()


options_df = options_df.sort_values(
    "timestamp"
)

spot_df = spot_df.sort_values(
    "timestamp"
)


merged = pd.merge_asof(

    options_df,

    spot_df,

    on="timestamp",

    direction="backward"

)


# ============================================================
# IV / GREEKS
# ============================================================

print()
print(
    "Calculating IV + Greeks..."
)


expiry_datetime = datetime.strptime(
    EXPIRY_DATE,
    "%Y-%m-%d"
).replace(
    hour=15,
    minute=30,
    tzinfo=IST
)


iv_values = []

delta_values = []

gamma_values = []

theta_values = []

vega_values = []

dte_values = []


for _, row in merged.iterrows():

    S = safe_float(
        row.get(
            "spot_close"
        )
    )

    K = safe_float(
        row.get(
            "strike"
        )
    )

    option_price = safe_float(
        row.get(
            "close"
        )
    )

    option_type = row[
        "option_type"
    ]

    timestamp = row[
        "timestamp"
    ]


    seconds = (
        expiry_datetime
        -
        timestamp
    ).total_seconds()


    dte = max(

        seconds
        /
        (
            24
            * 60
            * 60
        ),

        0.0

    )


    dte_values.append(
        dte
    )


    T = max(

        seconds
        /
        (
            DAYS_PER_YEAR
            * 24
            * 60
            * 60
        ),

        1.0 / DAYS_PER_YEAR

    )


    iv = solve_iv(

        option_price,

        S,

        K,

        T,

        RISK_FREE_RATE,

        option_type

    )


    if np.isfinite(iv):

        iv_percent = (
            iv * 100.0
        )

    else:

        iv_percent = np.nan


    iv_values.append(
        iv_percent
    )


    greeks = calculate_greeks(

        S,

        K,

        T,

        RISK_FREE_RATE,

        iv,

        option_type

    )


    delta_values.append(
        greeks["Delta"]
    )

    gamma_values.append(
        greeks["Gamma"]
    )

    theta_values.append(
        greeks["Theta"]
    )

    vega_values.append(
        greeks["Vega"]
    )


merged[
    "DTE"
] = dte_values


merged[
    "IV"
] = iv_values


merged[
    "Delta"
] = delta_values


merged[
    "Gamma"
] = gamma_values


merged[
    "Theta"
] = theta_values


merged[
    "Vega"
] = vega_values


# ============================================================
# DAILY VOLUME
#
# Sum all 5-minute candle volume for each option/date.
#
# This is the important part for comparing with daily
# NSE TOT_TRADED_QTY.
# ============================================================

print()
print(
    "Calculating daily traded quantity..."
)


merged[
    "Daily_Volume_RAW"
] = (

    merged
    .groupby(
        [
            "trade_date",
            "instrument_key"
        ]
    )[
        "volume_raw"
    ]
    .transform(
        "sum"
    )

)


merged[
    "Daily_Volume"
] = (

    merged[
        "Daily_Volume_RAW"
    ]
    /
    LOT_SIZE

)


# ============================================================
# CE / PE
# ============================================================

ce = merged[
    merged[
        "option_type"
    ] == "CE"
].copy()


pe = merged[
    merged[
        "option_type"
    ] == "PE"
].copy()


# ============================================================
# CE OI
#
# CE_OI_RAW = daily raw Upstox historical OI
# CE_OI     = CE_OI_RAW / LOT_SIZE
# CE_Change_OI = DAILY OI CHANGE, not 5-minute diff
#
# Historical OI from /market/oi is a daily snapshot.
# The same daily OI is therefore repeated on every 5-minute
# candle for that trading date.
# Change OI must be calculated from the daily OI table BEFORE
# it is merged into the 5-minute option candles.
# ============================================================

ce[
    "CE_OI"
] = (

    ce[
        "CE_OI_RAW"
    ]
    /
    LOT_SIZE

)


ce[
    "CE_Volume"
] = (

    ce[
        "volume_raw"
    ]
    /
    LOT_SIZE

)


ce[
    "CE_Daily_TOT_TRADED_QTY"
] = ce[
    "Daily_Volume_RAW"
]


ce[
    "CE_Daily_Volume"
] = ce[
    "Daily_Volume"
]


# ============================================================
# PE OI
# ============================================================

pe[
    "PE_OI"
] = (

    pe[
        "PE_OI_RAW"
    ]
    /
    LOT_SIZE

)


pe[
    "PE_Volume"
] = (

    pe[
        "volume_raw"
    ]
    /
    LOT_SIZE

)


pe[
    "PE_Daily_TOT_TRADED_QTY"
] = pe[
    "Daily_Volume_RAW"
]


pe[
    "PE_Daily_Volume"
] = pe[
    "Daily_Volume"
]


# ============================================================
# RENAME CE
# ============================================================

ce = ce.rename(

    columns={

        "open":
            "CE_Open",

        "high":
            "CE_High",

        "low":
            "CE_Low",

        "close":
            "CE_Close",

        "volume_raw":
            "CE_Volume_RAW",

        "CE_OI_RAW":
            "CE_OI_RAW",

        "CE_IV":
            "CE_IV",

        "IV":
            "CE_IV",

        "Delta":
            "CE_Delta",

        "Gamma":
            "CE_Gamma",

        "Theta":
            "CE_Theta",

        "Vega":
            "CE_Vega",

        "DTE":
            "CE_DTE"

    }

)


# ============================================================
# RENAME PE
# ============================================================

pe = pe.rename(

    columns={

        "open":
            "PE_Open",

        "high":
            "PE_High",

        "low":
            "PE_Low",

        "close":
            "PE_Close",

        "volume_raw":
            "PE_Volume_RAW",

        "IV":
            "PE_IV",

        "Delta":
            "PE_Delta",

        "Gamma":
            "PE_Gamma",

        "Theta":
            "PE_Theta",

        "Vega":
            "PE_Vega",

        "DTE":
            "PE_DTE"

    }

)


# ============================================================
# CE COLUMNS
# ============================================================

ce_columns = [

    "timestamp",

    "trade_date",

    "strike",

    "spot_open",

    "spot_high",

    "spot_low",

    "spot_close",

    "spot_volume",

    "spot_oi",

    "CE_Open",

    "CE_High",

    "CE_Low",

    "CE_Close",

    "CE_Volume_RAW",

    "CE_Volume",

    "CE_Daily_TOT_TRADED_QTY",

    "CE_Daily_Volume",

    "CE_OI_RAW",

    "CE_OI",

    "CE_Change_OI",

    "CE_IV",

    "CE_Delta",

    "CE_Gamma",

    "CE_Theta",

    "CE_Vega",

    "CE_DTE",

    "trading_symbol",

    "instrument_key"

]


ce_columns = [

    column

    for column
    in ce_columns

    if column in ce.columns

]


ce = ce[
    ce_columns
]


# ============================================================
# PE COLUMNS
# ============================================================

pe_columns = [

    "timestamp",

    "trade_date",

    "strike",

    "PE_Open",

    "PE_High",

    "PE_Low",

    "PE_Close",

    "PE_Volume_RAW",

    "PE_Volume",

    "PE_Daily_TOT_TRADED_QTY",

    "PE_Daily_Volume",

    "PE_OI_RAW",

    "PE_OI",

    "PE_Change_OI",

    "PE_IV",

    "PE_Delta",

    "PE_Gamma",

    "PE_Theta",

    "PE_Vega",

    "PE_DTE",

    "trading_symbol",

    "instrument_key"

]


pe_columns = [

    column

    for column
    in pe_columns

    if column in pe.columns

]


pe = pe[
    pe_columns
]


# ============================================================
# RENAME CONTRACT INFO
# ============================================================

ce = ce.rename(

    columns={

        "trading_symbol":
            "CE_Symbol",

        "instrument_key":
            "CE_Instrument_Key"

    }

)


pe = pe.rename(

    columns={

        "trading_symbol":
            "PE_Symbol",

        "instrument_key":
            "PE_Instrument_Key"

    }

)


# ============================================================
# MERGE CE + PE
# ============================================================

chain = pd.merge(

    ce,

    pe,

    on=[
        "timestamp",
        "strike"
    ],

    how="outer"

)


# ============================================================
# SORT
# ============================================================

chain = chain.sort_values(

    [
        "timestamp",
        "strike"
    ]

).reset_index(
    drop=True
)


# ============================================================
# CHANGE OI
#
# IMPORTANT FIX
#
# CE_OI and PE_OI are DAILY OI snapshots.
# They are repeated across all 5-minute candles for a date.
# Therefore a 5-minute diff would produce 0 on almost every row.
#
# We already calculated the DAILY change in oi_df above.
# Just carry that daily change into the final chain.
#
# CE_Change_OI = today's CE_OI - previous trading day's CE_OI
# PE_Change_OI = today's PE_OI - previous trading day's PE_OI
# ============================================================

chain = chain.sort_values(

    [
        "timestamp",
        "strike"
    ]

).reset_index(
    drop=True
)


# Safety: if the daily change columns were not created,
# create them as NaN instead of generating incorrect 5-minute diffs.

if "CE_Change_OI" not in chain.columns:

    chain[
        "CE_Change_OI"
    ] = np.nan


if "PE_Change_OI" not in chain.columns:

    chain[
        "PE_Change_OI"
    ] = np.nan


# ============================================================
# PCR
#
# PUT OI / CALL OI
# ============================================================

chain[
    "PCR_OI"
] = np.where(

    chain[
        "CE_OI"
    ].notna()
    &
    (
        chain[
            "CE_OI"
        ].abs() > 0
    ),

    chain[
        "PE_OI"
    ]
    /
    chain[
        "CE_OI"
    ],

    np.nan

)


# ============================================================
# PCR CHANGE OI
# ============================================================

chain[
    "PCR_Change_OI"
] = np.where(

    chain[
        "CE_Change_OI"
    ].notna()
    &
    (
        chain[
            "CE_Change_OI"
        ].abs() > 0
    ),

    chain[
        "PE_Change_OI"
    ]
    /
    chain[
        "CE_Change_OI"
    ],

    np.nan

)


# ============================================================
# TOTAL OI
# ============================================================

chain[
    "Total_OI"
] = (

    chain[
        "CE_OI"
    ].fillna(0)

    +

    chain[
        "PE_OI"
    ].fillna(0)

)


# ============================================================
# TOTAL RAW OI
# ============================================================

chain[
    "Total_OI_RAW"
] = (

    chain[
        "CE_OI_RAW"
    ].fillna(0)

    +

    chain[
        "PE_OI_RAW"
    ].fillna(0)

)


# ============================================================
# OI DIFFERENCE
# ============================================================

chain[
    "OI_Difference"
] = (

    chain[
        "PE_OI"
    ]

    -

    chain[
        "CE_OI"
    ]

)


# ============================================================
# IV DIFFERENCE
# ============================================================

chain[
    "IV_Difference"
] = (

    chain[
        "PE_IV"
    ]

    -

    chain[
        "CE_IV"
    ]

)


# ============================================================
# ATM DISTANCE
# ============================================================

chain[
    "ATM_Distance"
] = (

    chain[
        "strike"
    ]

    -

    chain[
        "spot_close"
    ]

)


# ============================================================
# MONEYNESS
# ============================================================

def get_moneyness(
    strike,
    spot
):

    if (
        not np.isfinite(strike)
        or
        not np.isfinite(spot)
    ):

        return ""


    if abs(
        strike - spot
    ) <= 25:

        return "ATM"


    if strike < spot:

        return "ITM"


    return "OTM"


chain[
    "Moneyness"
] = [

    get_moneyness(
        row["strike"],
        row["spot_close"]
    )

    for _, row
    in chain.iterrows()

]


# ============================================================
# FINAL SORT
# ============================================================

chain = chain.sort_values(

    [
        "timestamp",
        "strike"
    ]

).reset_index(
    drop=True
)


# ============================================================
# FINAL COLUMN ORDER
# ============================================================

final_columns = [

    "timestamp",

    "trade_date",

    # SPOT
    "spot_open",
    "spot_high",
    "spot_low",
    "spot_close",
    "spot_volume",
    "spot_oi",

    # CALL
    "CE_Open",
    "CE_High",
    "CE_Low",
    "CE_Close",

    "CE_Volume_RAW",
    "CE_Volume",

    "CE_Daily_TOT_TRADED_QTY",
    "CE_Daily_Volume",

    "CE_OI_RAW",
    "CE_OI",
    "CE_Change_OI",

    "CE_IV",
    "CE_Delta",
    "CE_Gamma",
    "CE_Theta",
    "CE_Vega",
    "CE_DTE",

    # STRIKE
    "strike",
    "Moneyness",
    "ATM_Distance",

    # OI
    "PCR_OI",
    "PCR_Change_OI",

    "Total_OI_RAW",
    "Total_OI",

    "OI_Difference",

    "IV_Difference",

    # PUT
    "PE_Open",
    "PE_High",
    "PE_Low",
    "PE_Close",

    "PE_Volume_RAW",
    "PE_Volume",

    "PE_Daily_TOT_TRADED_QTY",
    "PE_Daily_Volume",

    "PE_OI_RAW",
    "PE_OI",
    "PE_Change_OI",

    "PE_IV",
    "PE_Delta",
    "PE_Gamma",
    "PE_Theta",
    "PE_Vega",
    "PE_DTE"

]


final_columns = [

    column

    for column
    in final_columns

    if column in chain.columns

]


chain = chain[
    final_columns
]


# ============================================================
# CLEAN INF
# ============================================================

chain = chain.replace(

    [
        np.inf,
        -np.inf
    ],

    np.nan

)


spot_excel = spot_df.copy()

spot_excel = spot_excel.replace(

    [
        np.inf,
        -np.inf
    ],

    np.nan

)


contracts_excel = (
    selected_contracts
    .copy()
)


oi_excel = oi_df.copy()


# ============================================================
# EXCEL DATETIME
# ============================================================

def make_excel_datetime_naive(
    df,
    column
):

    if column not in df.columns:

        return


    dt = pd.to_datetime(
        df[column],
        errors="coerce"
    )


    try:

        if dt.dt.tz is not None:

            dt = (
                dt
                .dt
                .tz_localize(
                    None
                )
            )

    except Exception:

        pass


    df[column] = dt


make_excel_datetime_naive(
    chain,
    "timestamp"
)

make_excel_datetime_naive(
    spot_excel,
    "timestamp"
)


chain = chain.reset_index(
    drop=True
)

spot_excel = spot_excel.reset_index(
    drop=True
)


# ============================================================
# SUMMARY
# ============================================================

summary = pd.DataFrame({

    "Parameter": [

        "Underlying",

        "Underlying Key",

        "Expiry",

        "Historical OI Expiry Column",

        "Requested Start",

        "Requested End",

        "Effective End",

        "Interval",

        "Lot Size",

        "Latest Historical Spot",

        "ATM Strike",

        "Lower Strike",

        "Upper Strike",

        "Selected Strikes",

        "CE Contracts",

        "PE Contracts",

        "Raw Option Candles",

        "Historical OI Rows",

        "Final Chain Rows",

        "Risk Free Rate"

    ],


    "Value": [

        UNDERLYING_NAME,

        UNDERLYING_KEY,

        EXPIRY_DATE,

        "expiry",

        START_DATE,

        END_DATE,

        effective_end_date.strftime(
            "%Y-%m-%d"
        ),

        "5 minutes",

        LOT_SIZE,

        latest_spot,

        ATM_STRIKE,

        lower_strike,

        upper_strike,

        len(
            selected_strikes
        ),

        len(
            selected_contracts[
                selected_contracts[
                    "option_type"
                ] == "CE"
            ]
        ),

        len(
            selected_contracts[
                selected_contracts[
                    "option_type"
                ] == "PE"
            ]
        ),

        len(
            option_records
        ),

        len(
            oi_df
        ),

        len(
            chain
        ),

        RISK_FREE_RATE

    ]

})


# ============================================================
# COLUMN ORDER FOR OptionChain_AllStrikes
#
# Puts Date / identity / OI columns on the LEFT where they are
# easy to see, instead of buried at the far right of a wide
# sheet. Everything else keeps its original relative order.
# ============================================================

ALL_STRIKE_FRONT_COLUMNS = [
    "Date",
    "expiry",
    "underlying_spot",
    "strike",
    "Moneyness",

    "CE_OI_RAW",
    "CE_OI",
    "CE_Previous_OI",
    "CE_Change_OI",

    "PE_OI_RAW",
    "PE_OI",
    "PE_Previous_OI",
    "PE_Change_OI",

    "Total_OI",
    "OI_Difference",
    "PCR_OI",
    "PCR_Change_OI"
]


def reorder_all_strike_columns(df):

    front_cols = [
        c for c in ALL_STRIKE_FRONT_COLUMNS
        if c in df.columns
    ]

    remaining_cols = [
        c for c in df.columns
        if c not in front_cols
    ]

    return df[front_cols + remaining_cols]


# ============================================================
# FORCE ALL-STRIKE OI TO USE THE SAME HISTORICAL OI TABLE
# AS OPTIONCHAIN_5MIN
# ============================================================

def sync_all_strike_oi_with_historical(all_strike_df, historical_oi_df, target_date=None):

    # ------------------------------------------------------------
    # FALLBACK DATE
    #
    # Used whenever historical OI is unavailable, so the Date
    # column on OptionChain_AllStrikes is NEVER blank.
    # ------------------------------------------------------------

    fallback_date_string = datetime.now(IST).date().strftime("%Y-%m-%d")

    if all_strike_df is None or all_strike_df.empty:
        return all_strike_df

    if historical_oi_df is None or historical_oi_df.empty:

        all_strike_df = all_strike_df.copy()
        all_strike_df["Date"] = fallback_date_string
        all_strike_df["OI_Reference_Date"] = fallback_date_string
        all_strike_df["OI_Previous_Trade_Date"] = ""

        all_strike_df = reorder_all_strike_columns(all_strike_df)

        print()
        print("ALL-STRIKE OI SYNC SKIPPED")
        print("Reason: Historical_OI data not available.")
        print("Date column filled with today's date as fallback:", fallback_date_string)

        return all_strike_df

    hist = historical_oi_df.copy()

    # ------------------------------------------------------------
    # NORMALIZE DATE AND STRIKE
    # ------------------------------------------------------------
    hist["trade_date"] = pd.to_datetime(
        hist["trade_date"],
        errors="coerce"
    ).dt.date

    hist["strike"] = pd.to_numeric(
        hist["strike"],
        errors="coerce"
    ).round(2)

    hist = hist.dropna(
        subset=["trade_date", "strike"]
    ).copy()

    if hist.empty:

        all_strike_df = all_strike_df.copy()
        all_strike_df["Date"] = fallback_date_string
        all_strike_df["OI_Reference_Date"] = fallback_date_string
        all_strike_df["OI_Previous_Trade_Date"] = ""

        all_strike_df = reorder_all_strike_columns(all_strike_df)

        print()
        print("ALL-STRIKE OI SYNC SKIPPED")
        print("Reason: Historical_OI has no valid trade_date/strike rows.")
        print("Date column filled with today's date as fallback:", fallback_date_string)

        return all_strike_df

    # ------------------------------------------------------------
    # KEEP ONE ROW PER DATE + STRIKE
    # ------------------------------------------------------------
    hist = hist.sort_values(
        ["trade_date", "strike"]
    ).drop_duplicates(
        subset=["trade_date", "strike"],
        keep="last"
    ).reset_index(drop=True)

    # ------------------------------------------------------------
    # IMPORTANT
    #
    # If target_date is specified, use that date for OI.
    # Otherwise, use the latest date.
    #
    # CHANGE OI IS ALWAYS CALCULATED FROM RAW OI:
    #
    # CE_Change_OI = (Current CE_OI_RAW
    #                 - Previous CE_OI_RAW) / 65
    #
    # PE_Change_OI = (Current PE_OI_RAW
    #                 - Previous PE_OI_RAW) / 65
    # ------------------------------------------------------------

    all_dates = sorted(
        hist["trade_date"].dropna().unique()
    )

    # If target_date is specified, use it
    if target_date is not None:
        target_date = pd.to_datetime(target_date).date()
        if target_date in all_dates:
            latest_date = target_date
            # Find previous date
            idx = all_dates.index(latest_date)
            previous_date = all_dates[idx - 1] if idx > 0 else None
        else:
            # If target_date not in historical data, use latest
            print(f"Warning: target_date {target_date} not found in historical OI. Using latest date instead.")
            latest_date = all_dates[-1]
            previous_date = all_dates[-2] if len(all_dates) >= 2 else None
    else:
        latest_date = all_dates[-1]
        previous_date = all_dates[-2] if len(all_dates) >= 2 else None

    current = hist[
        hist["trade_date"] == latest_date
    ].copy()

    previous = pd.DataFrame()

    if previous_date is not None:
        previous = hist[
            hist["trade_date"] == previous_date
        ].copy()

    # ------------------------------------------------------------
    # CURRENT OI
    # ------------------------------------------------------------

    current = current[[
        c for c in [
            "strike",
            "CE_OI_RAW",
            "PE_OI_RAW"
        ]
        if c in current.columns
    ]].rename(columns={
        "CE_OI_RAW": "SYNC_CE_OI_RAW",
        "PE_OI_RAW": "SYNC_PE_OI_RAW"
    })

    # ------------------------------------------------------------
    # PREVIOUS OI
    # ------------------------------------------------------------

    if not previous.empty:

        previous = previous[[
            c for c in [
                "strike",
                "CE_OI_RAW",
                "PE_OI_RAW"
            ]
            if c in previous.columns
        ]].rename(columns={
            "CE_OI_RAW": "SYNC_CE_PREVIOUS_OI_RAW",
            "PE_OI_RAW": "SYNC_PE_PREVIOUS_OI_RAW"
        })

    else:

        previous = pd.DataFrame(columns=[
            "strike",
            "SYNC_CE_PREVIOUS_OI_RAW",
            "SYNC_PE_PREVIOUS_OI_RAW"
        ])

    # ------------------------------------------------------------
    # PREPARE ALL-STRIKE DATA
    # ------------------------------------------------------------

    result = all_strike_df.copy()

    result["strike"] = pd.to_numeric(
        result["strike"],
        errors="coerce"
    ).round(2)

    # Remove old OI columns before merging
    old_oi_columns = [
        "CE_OI_RAW",
        "CE_OI",
        "CE_Previous_OI_RAW",
        "CE_Previous_OI",
        "CE_Change_OI",
        "PE_OI_RAW",
        "PE_OI",
        "PE_Previous_OI_RAW",
        "PE_Previous_OI",
        "PE_Change_OI"
    ]

    result = result.drop(
        columns=[
            c for c in old_oi_columns
            if c in result.columns
        ],
        errors="ignore"
    )

    result = result.merge(
        current,
        on="strike",
        how="left"
    )

    result = result.merge(
        previous,
        on="strike",
        how="left"
    )

    # ------------------------------------------------------------
    # RAW OI (restored)
    # ------------------------------------------------------------

    result["CE_OI_RAW"] = result["SYNC_CE_OI_RAW"]
    result["PE_OI_RAW"] = result["SYNC_PE_OI_RAW"]

    # ------------------------------------------------------------
    # CURRENT OI / 65
    # ------------------------------------------------------------

    result["CE_OI"] = (
        result["SYNC_CE_OI_RAW"]
        / LOT_SIZE
    )

    result["PE_OI"] = (
        result["SYNC_PE_OI_RAW"]
        / LOT_SIZE
    )

    # ------------------------------------------------------------
    # PREVIOUS OI / 65
    # ------------------------------------------------------------

    result["CE_Previous_OI"] = (
        result["SYNC_CE_PREVIOUS_OI_RAW"]
        / LOT_SIZE
    )

    result["PE_Previous_OI"] = (
        result["SYNC_PE_PREVIOUS_OI_RAW"]
        / LOT_SIZE
    )

    # ------------------------------------------------------------
    # CHANGE OI
    # ------------------------------------------------------------

    result["CE_Change_OI"] = (
        result["CE_OI"]
        - result["CE_Previous_OI"]
    )

    result["PE_Change_OI"] = (
        result["PE_OI"]
        - result["PE_Previous_OI"]
    )

    # ------------------------------------------------------------
    # REFERENCE DATE
    # ------------------------------------------------------------

    result["trade_date"] = latest_date

    result["Date"] = latest_date.strftime("%Y-%m-%d")

    result["OI_Reference_Date"] = latest_date.strftime(
        "%Y-%m-%d"
    )

    result["OI_Previous_Trade_Date"] = (
        previous_date.strftime("%Y-%m-%d")
        if previous_date is not None
        else ""
    )

    # ------------------------------------------------------------
    # OI-DERIVED VALUES
    # ------------------------------------------------------------

    result["PCR_OI"] = np.where(
        result["CE_OI"].notna()
        & (result["CE_OI"].abs() > 0),
        result["PE_OI"] / result["CE_OI"],
        np.nan
    )

    result["PCR_Change_OI"] = np.where(
        result["CE_Change_OI"].notna()
        & (result["CE_Change_OI"].abs() > 0),
        result["PE_Change_OI"]
        / result["CE_Change_OI"],
        np.nan
    )

    result["Total_OI"] = (
        result["CE_OI"].fillna(0)
        + result["PE_OI"].fillna(0)
    )

    result["OI_Difference"] = (
        result["PE_OI"]
        - result["CE_OI"]
    )

    # ------------------------------------------------------------
    # CLEAN TEMPORARY COLUMNS
    # ------------------------------------------------------------

    result = result.drop(
        columns=[
            "SYNC_CE_OI_RAW",
            "SYNC_PE_OI_RAW",
            "SYNC_CE_PREVIOUS_OI_RAW",
            "SYNC_PE_PREVIOUS_OI_RAW"
        ],
        errors="ignore"
    )

    # ------------------------------------------------------------
    # REORDER COLUMNS
    # ------------------------------------------------------------

    result = reorder_all_strike_columns(result)

    print()
    print("ALL-STRIKE OI SYNC COMPLETE")
    print("OI reference date:", latest_date)
    print("Previous OI date :", previous_date)
    print("OI formula       : CURRENT RAW / 65")
    print("CHANGE formula   : (CURRENT RAW - PREVIOUS RAW) / 65")

    # ------------------------------------------------------------
    # DEBUG CHECK FOR 24200
    # ------------------------------------------------------------

    check_24200 = result[
        result["strike"] == 24200
    ]

    if not check_24200.empty:
        r = check_24200.iloc[0]
        print()
        print("24200 CE OI CHECK")
        print("Current CE OI     :", r.get("CE_OI"))
        print("Previous CE OI    :", r.get("CE_Previous_OI"))
        print("CE Change OI      :", r.get("CE_Change_OI"))
        print("Current trade date:", r.get("trade_date"))

    return result


# ============================================================
# NSE-WEBSITE-STYLE OPTION CHAIN
#
# Lays data out exactly like nseindia.com/option-chain:
#
# CALLS (left -> right):
#   OI, CHNG IN OI, VOLUME, IV, LTP, CHNG, BID QTY, BID, ASK, ASK QTY
#
# STRIKE (center)
#
# PUTS (left -> right):
#   BID QTY, BID, ASK, ASK QTY, CHNG, LTP, IV, VOLUME, CHNG IN OI, OI
#
# Several NSE column names repeat on both sides (OI, BID, etc.),
# so this is built as a plain header list + row matrix instead
# of a pandas DataFrame (pandas can't hold duplicate column
# names cleanly). The Excel-writing step writes these directly
# with openpyxl.
# ============================================================

NSE_CHAIN_HEADERS = [
    "OI", "CHNG IN OI", "VOLUME", "IV", "LTP", "CHNG",
    "BID QTY", "BID", "ASK", "ASK QTY",

    "STRIKE",

    "BID QTY", "BID", "ASK", "ASK QTY", "CHNG", "LTP", "IV",
    "VOLUME", "CHNG IN OI", "OI"
]


def build_nse_style_full(all_strike_df, is_historical=False):
    """
    Build NSE-style rows from all-strike data.

    If is_historical=True (a past date, built from 5-min candles):
    - CHNG   = this date's close vs the PREVIOUS TRADING DAY's
               close (matches NSE's own "Chng" definition), using
               CE_Prev_Close/PE_Prev_Close looked up in
               build_all_strike_chain_from_history().
    - BID / ASK / BID QTY / ASK QTY are left blank. There is no
      historical order-book depth available anywhere (not from
      Upstox, not from NSE's public site or Bhavcopy) for a past
      date -- only a live snapshot exists -- so these are never
      approximated from OHLC/Volume, which produced numbers that
      could never actually match NSE.
    """
    if all_strike_df is None or all_strike_df.empty:
        return [], ""

    df = all_strike_df.copy()

    df["strike"] = pd.to_numeric(df["strike"], errors="coerce")
    df = df.dropna(subset=["strike"]).sort_values("strike").reset_index(drop=True)

    date_label = ""
    if "Date" in df.columns and not df["Date"].empty:
        date_label = str(df["Date"].iloc[0])

    def safe_value(value):
        """Safely convert value to float or return nan"""
        if value is None:
            return np.nan
        try:
            return float(value)
        except (ValueError, TypeError):
            return np.nan

    def g(row, col):
        value = row.get(col)
        if value is None:
            return ""
        try:
            if isinstance(value, float) and np.isnan(value):
                return ""
            if isinstance(value, str) and value == "":
                return ""
            return value
        except Exception:
            return ""

    def get_bid_ask(row, option_type):
        """
        Bid/Ask/Bid Qty/Ask Qty are live order-book fields.
        Upstox (and NSE's own public site/Bhavcopy) never exposes
        historical order-book depth for a past date -- only a live
        snapshot -- so for historical rows these are genuinely
        blank rather than approximated from candle High/Low/Volume
        (that was giving numbers that could never match NSE, since
        High/Low/Volume aren't bid/ask at all).
        """
        if is_historical:
            return "", "", "", ""

        if option_type == "CE":
            bid = safe_value(row.get("CE_Bid"))
            ask = safe_value(row.get("CE_Ask"))
            bid_qty = g(row, "CE_Bid_Qty")
            ask_qty = g(row, "CE_Ask_Qty")
        else:
            bid = safe_value(row.get("PE_Bid"))
            ask = safe_value(row.get("PE_Ask"))
            bid_qty = g(row, "PE_Bid_Qty")
            ask_qty = g(row, "PE_Ask_Qty")

        if np.isfinite(bid) and np.isfinite(ask):
            return bid_qty, bid, ask, ask_qty

        close = safe_value(row.get("CE_Close" if option_type == "CE" else "PE_Close"))

        if np.isfinite(close):
            spread = abs(close) * 0.001 if abs(close) > 0 else 0.5
            return "", close - spread, close + spread, ""

        return "", "", "", ""

    def get_change(row, option_type):
        """
        NSE's "Chng" column is today's price vs the PREVIOUS
        TRADING DAY's close -- not the same candle's own
        open-to-close move (that was the bug: it was showing an
        intraday move, not the day-over-day change NSE shows).
        """
        if is_historical:

            if option_type == "CE":
                close = safe_value(row.get("CE_Close"))
                prev_close = safe_value(row.get("CE_Prev_Close"))
            else:
                close = safe_value(row.get("PE_Close"))
                prev_close = safe_value(row.get("PE_Prev_Close"))

            if np.isfinite(close) and np.isfinite(prev_close):
                return close - prev_close

            return ""

        # For live data, use LTP - Close (Close = previous day's close
        # as reported by Upstox's live option-chain endpoint).
        if option_type == "CE":
            ltp = safe_value(row.get("CE_LTP"))
            close = safe_value(row.get("CE_Close"))
        else:
            ltp = safe_value(row.get("PE_LTP"))
            close = safe_value(row.get("PE_Close"))

        if np.isfinite(ltp) and np.isfinite(close):
            return ltp - close

        return ""

    rows = []

    for _, row in df.iterrows():
        
        # Get CE bid/ask
        ce_bid_qty, ce_bid, ce_ask, ce_ask_qty = get_bid_ask(row, "CE")
        pe_bid_qty, pe_bid, pe_ask, pe_ask_qty = get_bid_ask(row, "PE")
        
        # Get change
        ce_chng = get_change(row, "CE")
        pe_chng = get_change(row, "PE")

        rows.append([
            # ---- CALLS ----
            g(row, "CE_OI"),
            g(row, "CE_Change_OI"),
            g(row, "CE_Volume"),
            g(row, "CE_IV"),
            g(row, "CE_LTP"),
            ce_chng if ce_chng != "" else "",
            ce_bid_qty,
            ce_bid,
            ce_ask,
            ce_ask_qty,

            # ---- STRIKE ----
            g(row, "strike"),

            # ---- PUTS ----
            pe_bid_qty,
            pe_bid,
            pe_ask,
            pe_ask_qty,
            pe_chng if pe_chng != "" else "",
            g(row, "PE_LTP"),
            g(row, "PE_IV"),
            g(row, "PE_Volume"),
            g(row, "PE_Change_OI"),
            g(row, "PE_OI"),
        ])

    return rows, date_label


def write_nse_style_sheet(wb, sheet_name, title_text, rows, atm_strike=None):

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill(
        start_color="312B69", end_color="312B69", fill_type="solid"
    )
    band_fill = PatternFill(
        start_color="312B69", end_color="312B69", fill_type="solid"
    )
    strike_fill = PatternFill(
        start_color="EDEDED", end_color="EDEDED", fill_type="solid"
    )
    atm_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    white_bold = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ---- Title row ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(NSE_CHAIN_HEADERS))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="left")

    # ---- CALLS / PUTS band ----
    band_row = 2
    ws.merge_cells(start_row=band_row, start_column=1, end_row=band_row, end_column=10)
    calls_cell = ws.cell(row=band_row, column=1, value="CALLS")
    calls_cell.font = white_bold
    calls_cell.alignment = center
    calls_cell.fill = band_fill

    strike_band_cell = ws.cell(row=band_row, column=11, value="")
    strike_band_cell.fill = band_fill

    ws.merge_cells(start_row=band_row, start_column=12, end_row=band_row, end_column=21)
    puts_cell = ws.cell(row=band_row, column=12, value="PUTS")
    puts_cell.font = white_bold
    puts_cell.alignment = center
    puts_cell.fill = band_fill

    # ---- Header row ----
    header_row = 3
    for col_idx, header in enumerate(NSE_CHAIN_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = white_bold
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border

    # ---- Data rows ----
    data_start = header_row + 1

    for r_offset, row_values in enumerate(rows):

        r = data_start + r_offset
        strike_value = row_values[10]

        is_atm = (
            atm_strike is not None
            and strike_value != ""
            and float(strike_value) == float(atm_strike)
        )

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            if col_idx == 11:
                cell.fill = atm_fill if is_atm else strike_fill
                cell.font = Font(bold=True)
            elif is_atm:
                cell.fill = atm_fill

    # ---- Column widths ----
    for col_idx in range(1, len(NSE_CHAIN_HEADERS) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 11 if col_idx != 11 else 12

    ws.freeze_panes = "A4"


# ============================================================
# ALL-STRIKE OPTION CHAIN + SINGLE EXPIRY LIST
# ============================================================

print()
print("=" * 80)
print("DOWNLOADING ALL-STRIKE OPTION CHAIN")
print("=" * 80)
print()


def build_all_strike_chain_from_history(chain_df, expiry_date_str, target_date=None):

    # ------------------------------------------------------------
    # PAST-EXPIRY FALLBACK OR SPECIFIC DATE SNAPSHOT.
    #
    # There is no live snapshot for an expired contract (Upstox's
    # /option/chain endpoint only covers active/upcoming expiries),
    # so this reconstructs an all-strike "snapshot" from the
    # already-downloaded 5-min historical chain instead.
    #
    # If target_date is specified, take the LAST candle on that
    # specific date. Otherwise, take the global last candle.
    #
    # Bid/Ask/Bid Qty/Ask Qty/POP/Previous OI have no historical
    # equivalent (they're live orderbook-only fields), so those
    # stay blank -- same for CE/PE Instrument Key.
    # ------------------------------------------------------------

    if chain_df is None or chain_df.empty:
        return pd.DataFrame()

    df = chain_df.copy()

    if "strike" not in df.columns:
        return pd.DataFrame()

    df = df.dropna(subset=["strike"])

    if df.empty:
        return pd.DataFrame()

    # If target_date is specified, filter to that date first
    prev_close_lookup = {}

    if target_date is not None:
        # Check if trade_date exists, if not create it from timestamp
        if "trade_date" not in df.columns:
            if "timestamp" in df.columns:
                df["trade_date"] = pd.to_datetime(df["timestamp"]).dt.date
            else:
                return pd.DataFrame()
        else:
            df["trade_date"] = pd.to_datetime(df["trade_date"]).dt.date

        # ------------------------------------------------------------
        # PREVIOUS TRADING DAY'S CLOSE (for a correct CHNG later).
        #
        # NSE's "Chng" column is today's price vs the PREVIOUS
        # TRADING DAY's close -- not vs the same candle's own open.
        # Look that up here, before filtering df down to target_date,
        # so build_nse_style_full can compute CHNG correctly instead
        # of the wrong close-minus-open-of-same-candle approximation.
        # ------------------------------------------------------------

        prev_df = df[df["trade_date"] < target_date]

        if not prev_df.empty:

            prev_df = prev_df.sort_values(["strike", "timestamp"])

            prev_latest = prev_df.groupby(
                "strike", as_index=False
            ).tail(1)

            for _, prow in prev_latest.iterrows():
                prev_close_lookup[prow["strike"]] = (
                    prow.get("CE_Close"),
                    prow.get("PE_Close"),
                )

        df = df[df["trade_date"] == target_date]
        if df.empty:
            return pd.DataFrame()

    df = df.sort_values(["strike", "timestamp"])

    # Get the last candle for each strike
    latest = df.groupby("strike", as_index=False).tail(1).reset_index(drop=True)

    records = []

    for _, row in latest.iterrows():

        strike = row.get("strike")
        spot_close = row.get("spot_close")

        ce_close = row.get("CE_Close")
        pe_close = row.get("PE_Close")
        
        # Get OHLC data
        ce_open = row.get("CE_Open")
        ce_high = row.get("CE_High")
        ce_low = row.get("CE_Low")
        pe_open = row.get("PE_Open")
        pe_high = row.get("PE_High")
        pe_low = row.get("PE_Low")

        moneyness = ""
        if np.isfinite(safe_float(strike)) and np.isfinite(safe_float(spot_close)):
            moneyness = get_moneyness(safe_float(strike), safe_float(spot_close))

        ce_change = row.get("CE_Change_OI")
        pe_change = row.get("PE_Change_OI")
        
        if ce_change is None or np.isnan(ce_change):
            ce_change = np.nan
        
        if pe_change is None or np.isnan(pe_change):
            pe_change = np.nan

        ce_prev_close, pe_prev_close = prev_close_lookup.get(
            strike, (np.nan, np.nan)
        )

        records.append({
            "expiry": expiry_date_str,
            "underlying_spot": spot_close,
            "strike": strike,
            "Moneyness": moneyness,

            # CE fields - include OHLC
            "CE_Instrument_Key": row.get("CE_Instrument_Key", ""),
            "CE_Open": ce_open,
            "CE_High": ce_high,
            "CE_Low": ce_low,
            "CE_LTP": ce_close,
            "CE_Close": ce_close,
            "CE_Prev_Close": ce_prev_close,
            "CE_Volume_RAW": row.get("CE_Volume_RAW"),
            "CE_Volume": row.get("CE_Volume"),
            "CE_OI_RAW": row.get("CE_OI_RAW"),
            "CE_OI": row.get("CE_OI"),
            "CE_Previous_OI_RAW": np.nan,
            "CE_Previous_OI": np.nan,
            "CE_Change_OI": ce_change,
            # Bid/Ask calculated from OHLC in build_nse_style_full
            "CE_Bid": np.nan,
            "CE_Bid_Qty": np.nan,
            "CE_Ask": np.nan,
            "CE_Ask_Qty": np.nan,
            "CE_IV": row.get("CE_IV"),
            "CE_Delta": row.get("CE_Delta"),
            "CE_Gamma": row.get("CE_Gamma"),
            "CE_Theta": row.get("CE_Theta"),
            "CE_Vega": row.get("CE_Vega"),
            "CE_POP": np.nan,

            "PCR_OI": row.get("PCR_OI"),

            # PE fields - include OHLC
            "PE_Instrument_Key": row.get("PE_Instrument_Key", ""),
            "PE_Open": pe_open,
            "PE_High": pe_high,
            "PE_Low": pe_low,
            "PE_LTP": pe_close,
            "PE_Close": pe_close,
            "PE_Prev_Close": pe_prev_close,
            "PE_Volume_RAW": row.get("PE_Volume_RAW"),
            "PE_Volume": row.get("PE_Volume"),
            "PE_OI_RAW": row.get("PE_OI_RAW"),
            "PE_OI": row.get("PE_OI"),
            "PE_Previous_OI_RAW": np.nan,
            "PE_Previous_OI": np.nan,
            "PE_Change_OI": pe_change,
            "PE_Bid": np.nan,
            "PE_Bid_Qty": np.nan,
            "PE_Ask": np.nan,
            "PE_Ask_Qty": np.nan,
            "PE_IV": row.get("PE_IV"),
            "PE_Delta": row.get("PE_Delta"),
            "PE_Gamma": row.get("PE_Gamma"),
            "PE_Theta": row.get("PE_Theta"),
            "PE_Vega": row.get("PE_Vega"),
            "PE_POP": np.nan,
        })

    out = pd.DataFrame(records)

    if not out.empty:
        out = out.sort_values("strike").reset_index(drop=True)
        out["Total_OI"] = out["CE_OI"].fillna(0) + out["PE_OI"].fillna(0)
        out["OI_Difference"] = out["PE_OI"] - out["CE_OI"]
        out["PCR_Change_OI"] = np.where(
            out["CE_Change_OI"].notna() & (out["CE_Change_OI"].abs() > 0),
            out["PE_Change_OI"] / out["CE_Change_OI"],
            np.nan
        )

    return out


if IS_PAST_EXPIRY:

    print("Past expiry: no live snapshot exists, building all-strike chain from historical 5-min data instead.")

    all_strike_chain = build_all_strike_chain_from_history(
        chain,
        EXPIRY_DATE
    )

else:

    all_strike_chain = build_all_strike_chain(EXPIRY_DATE)

all_strike_chain = sync_all_strike_oi_with_historical(
    all_strike_chain,
    oi_df,
    target_date=None  # Use latest for today's sheet
)

print(
    "All-strike rows:",
    len(all_strike_chain)
)


# ============================================================
# NSE-STYLE OPTION CHAIN -- BUILD ROWS (used at Excel-write time)
# ============================================================

nse_style_today_rows, nse_style_today_date = build_nse_style_full(
    all_strike_chain,
    is_historical=False
)

nse_style_atm_strike = None

if not all_strike_chain.empty and "Moneyness" in all_strike_chain.columns:

    atm_rows = all_strike_chain[
        all_strike_chain["Moneyness"] == "ATM"
    ]

    if not atm_rows.empty:
        nse_style_atm_strike = float(atm_rows.iloc[0]["strike"])


# Build NSE-style sheets for extra dates using historical data
nse_style_extra_sheets = []

for date_str in EXTRA_OI_DATES:

    try:
        parsed_date = datetime.strptime(
            date_str, "%Y-%m-%d"
        ).date()
    except Exception:
        print("Skipping invalid EXTRA_OI_DATES entry:", date_str)
        continue

    # Build all-strike chain for this specific date from historical data
    date_all_strike = build_all_strike_chain_from_history(
        chain,
        EXPIRY_DATE,
        target_date=parsed_date
    )

    if date_all_strike is not None and not date_all_strike.empty:
        print(f"Built all-strike chain for {date_str}, rows: {len(date_all_strike)}")
        
        # Sync OI with historical data - PASS THE TARGET DATE
        try:
            date_all_strike = sync_all_strike_oi_with_historical(
                date_all_strike,
                oi_df,
                target_date=parsed_date  # Use specific date for OI
            )
        except Exception as e:
            print(f"Warning: OI sync failed for {date_str}: {e}")
        
        # Now build NSE-style rows from this date's data - pass is_historical=True
        try:
            extra_rows, extra_date_label = build_nse_style_full(
                date_all_strike, 
                is_historical=True
            )
            
            if extra_rows:
                nse_style_extra_sheets.append(
                    (date_str, extra_rows, extra_date_label)
                )
                print(f"NSE-style rows built for {date_str}: {len(extra_rows)}")
            else:
                print(f"No NSE-style rows for {date_str}")
        except Exception as e:
            print(f"Error building NSE-style rows for {date_str}: {e}")
    else:
        print(f"No historical data found for {date_str}")


print()
print("=" * 80)
print("DOWNLOADING EXPIRY LIST")
print("=" * 80)
print()

expiry_df, active_expiries, past_expiries = build_expiry_dataframe()

print(
    "Expiry list rows:",
    len(expiry_df)
)


# ============================================================
# SQLITE DATABASE
# ============================================================

print()
print("=" * 80)
print("SAVING DATA TO SQLITE DATABASE")
print("=" * 80)
print()

try:
    db = sqlite3.connect(DB_FILE)

    chain.to_sql(
        "optionchain_5min",
        db,
        if_exists="replace",
        index=False
    )

    spot_excel.to_sql(
        "nifty_spot_5min",
        db,
        if_exists="replace",
        index=False
    )

    contracts_excel.to_sql(
        "contracts_selected",
        db,
        if_exists="replace",
        index=False
    )

    oi_excel.to_sql(
        "historical_oi",
        db,
        if_exists="replace",
        index=False
    )

    all_strike_chain.to_sql(
        "optionchain_all_strikes",
        db,
        if_exists="replace",
        index=False
    )

    expiry_df.to_sql(
        "expiry_list",
        db,
        if_exists="replace",
        index=False
    )

    # Remove old separate expiry tables from previous script versions.
    db.execute("DROP TABLE IF EXISTS coming_expiries")
    db.execute("DROP TABLE IF EXISTS past_expiries")

    db.commit()
    db.close()

    print("Database saved:")
    print(DB_FILE)

except Exception as error:
    print()
    print("DATABASE ERROR")
    print(error)


# ============================================================
# WRITE EXCEL
# ============================================================

print()
print("=" * 80)
print("WRITING EXCEL")
print("=" * 80)
print()


try:

    with pd.ExcelWriter(

        OUTPUT_FILE,

        engine="openpyxl"

    ) as writer:

        chain.to_excel(

            writer,

            sheet_name="OptionChain_5Min",

            index=False

        )


        spot_excel.to_excel(

            writer,

            sheet_name="NIFTY_Spot_5Min",

            index=False

        )


        contracts_excel.to_excel(

            writer,

            sheet_name="Contracts",

            index=False

        )


        oi_excel.to_excel(

            writer,

            sheet_name="Historical_OI",

            index=False

        )


        all_strike_chain.to_excel(

            writer,

            sheet_name="OptionChain_AllStrikes",

            index=False

        )


        expiry_df.to_excel(

            writer,

            sheet_name="Expiry_List",

            index=False

        )


        summary.to_excel(

            writer,

            sheet_name="Summary",

            index=False

        )


except Exception as error:

    print()
    print("=" * 80)
    print("EXCEL ERROR")
    print("=" * 80)
    print()

    print(
        error
    )

    traceback.print_exc()

    raise SystemExit


# ============================================================
# FORMAT EXCEL
# ============================================================

print(
    "Formatting Excel..."
)


try:

    from openpyxl import load_workbook

    from openpyxl.styles import Font

    from openpyxl.styles import Alignment

    from openpyxl.utils import get_column_letter


    wb = load_workbook(
        OUTPUT_FILE
    )


    for ws in wb.worksheets:

        ws.freeze_panes = "A2"


        if ws.max_row > 1:

            ws.auto_filter.ref = (
                ws.dimensions
            )


        for cell in ws[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )


        for col_idx in range(

            1,

            ws.max_column + 1

        ):

            letter = get_column_letter(
                col_idx
            )


            max_length = 0


            for cell in ws[letter]:

                try:

                    length = len(
                        str(
                            cell.value
                        )
                    )


                    if length > max_length:

                        max_length = length

                except Exception:

                    pass


            ws.column_dimensions[
                letter
            ].width = min(

                max(
                    max_length + 2,
                    10
                ),

                32

            )


    # ========================================================
    # NSE-WEBSITE-STYLE OPTION CHAIN SHEET(S)
    #
    # Isolated in its own try/except: if this fails for any
    # reason, the rest of the formatting (and the final
    # wb.save()) still goes ahead, and the real error prints
    # with a full traceback instead of being swallowed by the
    # generic "Excel formatting warning" at the bottom.
    # ========================================================

    try:

        if nse_style_today_rows:

            write_nse_style_sheet(
                wb,
                "OptionChain_NSE_Style",
                "NIFTY Option Chain  |  Expiry: " + str(EXPIRY_DATE)
                + "  |  As on: " + str(nse_style_today_date),
                nse_style_today_rows,
                atm_strike=nse_style_atm_strike
            )

            print("NSE-style sheet written: OptionChain_NSE_Style")
            print("Rows written:", len(nse_style_today_rows))

        else:

            print("NSE-style sheet SKIPPED: no all-strike rows to write.")

        for date_str, extra_rows, date_label in nse_style_extra_sheets:

            write_nse_style_sheet(
                wb,
                "OptionChain_NSE_" + date_str,
                "NIFTY Option Chain  |  Expiry: " + str(EXPIRY_DATE)
                + "  |  As on: " + str(date_label),
                extra_rows,
                atm_strike=None  # No ATM highlighting for past dates
            )

            print("NSE-style sheet written: OptionChain_NSE_" + date_str)

    except Exception as nse_style_error:

        print()
        print("=" * 80)
        print("NSE-STYLE SHEET ERROR (formatting continues without it)")
        print("=" * 80)
        print(nse_style_error)
        traceback.print_exc()


    # ========================================================
    # MAIN SHEET
    # ========================================================

    ws = wb[
        "OptionChain_5Min"
    ]


    header_map = {

        cell.value:
            cell.column

        for cell in ws[1]

    }


    # ========================================================
    # DATETIME
    # ========================================================

    for column_name in [

        "timestamp"

    ]:

        if column_name not in header_map:

            continue


        col = get_column_letter(

            header_map[
                column_name
            ]

        )


        for cell in ws[col][1:]:

            cell.number_format = (
                "yyyy-mm-dd hh:mm:ss"
            )


    # ========================================================
    # DECIMAL
    # ========================================================

    decimal_columns = [

        "spot_open",
        "spot_high",
        "spot_low",
        "spot_close",

        "CE_Open",
        "CE_High",
        "CE_Low",
        "CE_Close",

        "PE_Open",
        "PE_High",
        "PE_Low",
        "PE_Close",

        "CE_IV",
        "PE_IV",

        "CE_Delta",
        "PE_Delta",

        "CE_Gamma",
        "PE_Gamma",

        "CE_Theta",
        "PE_Theta",

        "CE_Vega",
        "PE_Vega",

        "PCR_OI",
        "PCR_Change_OI",

        "IV_Difference",

        "ATM_Distance"

    ]


    for column_name in decimal_columns:

        if column_name not in header_map:

            continue


        col = get_column_letter(

            header_map[
                column_name
            ]

        )


        for cell in ws[col][1:]:

            cell.number_format = (
                "0.0000"
            )


    # ========================================================
    # INTEGER
    # ========================================================

    integer_columns = [

        "spot_volume",
        "spot_oi",

        "CE_Volume_RAW",
        "CE_Volume",

        "CE_Daily_TOT_TRADED_QTY",
        "CE_Daily_Volume",

        "CE_OI_RAW",
        "CE_OI",
        "CE_Change_OI",

        "PE_Volume_RAW",
        "PE_Volume",

        "PE_Daily_TOT_TRADED_QTY",
        "PE_Daily_Volume",

        "PE_OI_RAW",
        "PE_OI",
        "PE_Change_OI",

        "Total_OI_RAW",
        "Total_OI",

        "OI_Difference"

    ]


    for column_name in integer_columns:

        if column_name not in header_map:

            continue


        col = get_column_letter(

            header_map[
                column_name
            ]

        )


        for cell in ws[col][1:]:

            cell.number_format = (
                "#,##0"
            )


    # ========================================================
    # STRIKE
    # ========================================================

    if "strike" in header_map:

        col = get_column_letter(

            header_map[
                "strike"
            ]

        )


        for cell in ws[col][1:]:

            cell.number_format = (
                "0"
            )


    # ========================================================
    # HISTORICAL OI SHEET
    # ========================================================

    if "Historical_OI" in wb.sheetnames:

        ws_oi = wb[
            "Historical_OI"
        ]


        oi_header_map = {

            cell.value:
                cell.column

            for cell in ws_oi[1]

        }


        for column_name in [

            "CE_OI_RAW",
            "PE_OI_RAW"

        ]:

            if column_name not in oi_header_map:

                continue


            col = get_column_letter(

                oi_header_map[
                    column_name
                ]

            )


            for cell in ws_oi[col][1:]:

                cell.number_format = (
                    "#,##0"
                )


    # ========================================================
    # ALL-STRIKE / EXPIRY SHEETS
    # ========================================================

    for sheet_name in [
        "OptionChain_AllStrikes",
        "Expiry_List"
    ]:

        if sheet_name not in wb.sheetnames:
            continue

        ws_extra = wb[sheet_name]
        extra_header = {
            cell.value: cell.column
            for cell in ws_extra[1]
        }

        for name in [
            "strike",
            "underlying_spot",
            "CE_LTP",
            "CE_Close",
            "PE_LTP",
            "PE_Close",
            "CE_IV",
            "PE_IV",
            "CE_Delta",
            "PE_Delta",
            "CE_Gamma",
            "PE_Gamma",
            "CE_Theta",
            "PE_Theta",
            "CE_Vega",
            "PE_Vega",
            "PCR_OI",
            "PCR_Change_OI"
        ]:
            if name not in extra_header:
                continue
            col = get_column_letter(extra_header[name])
            for cell in ws_extra[col][1:]:
                cell.number_format = "0.0000"

        for name in [
            "CE_Volume_RAW",
            "CE_Volume",
            "CE_OI_RAW",
            "CE_OI",
            "CE_Previous_OI_RAW",
            "CE_Previous_OI",
            "CE_Change_OI",
            "CE_Bid_Qty",
            "CE_Ask_Qty",
            "PE_Volume_RAW",
            "PE_Volume",
            "PE_OI_RAW",
            "PE_OI",
            "PE_Previous_OI_RAW",
            "PE_Previous_OI",
            "PE_Change_OI",
            "PE_Bid_Qty",
            "PE_Ask_Qty",
            "Total_OI",
            "OI_Difference"
        ]:
            if name not in extra_header:
                continue
            col = get_column_letter(extra_header[name])
            for cell in ws_extra[col][1:]:
                cell.number_format = "#,##0"


    wb.save(
        OUTPUT_FILE
    )

    print()
    print("Workbook saved with sheets:", wb.sheetnames)


except Exception as error:

    print()
    print(
        "Excel formatting warning:"
    )

    print(
        error
    )

    traceback.print_exc()


# ============================================================
# TEST 24200
# ============================================================

print()
print("=" * 80)
print("OI / VOLUME CHECK")
print("=" * 80)
print()


test_rows = chain[
    chain[
        "strike"
    ] == 24200
]


if not test_rows.empty:

    test_row = test_rows.iloc[-1]


    print(
        "Strike:",
        test_row["strike"]
    )

    print()

    print(
        "CE OI RAW:",
        test_row.get(
            "CE_OI_RAW"
        )
    )

    print(
        "CE OI / 65:",
        test_row.get(
            "CE_OI"
        )
    )

    print(
        "CE Change OI:",
        test_row.get(
            "CE_Change_OI"
        )
    )

    print()

    print(
        "PE OI RAW:",
        test_row.get(
            "PE_OI_RAW"
        )
    )

    print(
        "PE OI / 65:",
        test_row.get(
            "PE_OI"
        )
    )

    print(
        "PE Change OI:",
        test_row.get(
            "PE_Change_OI"
        )
    )

    print()

    print(
        "CE 5-min Volume RAW:",
        test_row.get(
            "CE_Volume_RAW"
        )
    )

    print(
        "CE 5-min Volume / 65:",
        test_row.get(
            "CE_Volume"
        )
    )

    print()

    print(
        "CE Daily TOT TRADED QTY:",
        test_row.get(
            "CE_Daily_TOT_TRADED_QTY"
        )
    )

    print(
        "CE Daily Volume / 65:",
        test_row.get(
            "CE_Daily_Volume"
        )
    )

    print()

    print(
        "PE Daily TOT TRADED QTY:",
        test_row.get(
            "PE_Daily_TOT_TRADED_QTY"
        )
    )

    print(
        "PE Daily Volume / 65:",
        test_row.get(
            "PE_Daily_Volume"
        )
    )

else:

    print(
        "24200 strike not found."
    )


# ============================================================
# FINAL
# ============================================================

print()
print("=" * 80)
print("SUCCESS")
print("=" * 80)
print()

print(
    "Excel file:"
)

print(
    OUTPUT_FILE
)

print()
print("Database:")
print(DB_FILE)

print()

print(
    "Final rows:",
    len(chain)
)

print(
    "Final columns:",
    len(chain.columns)
)

print()

print(
    "Sheets:"
)

print(
    "1. OptionChain_5Min"
)

print(
    "2. NIFTY_Spot_5Min"
)

print(
    "3. Contracts"
)

print(
    "4. Historical_OI"
)

print(
    "5. OptionChain_AllStrikes"
)

print(
    "6. Expiry_List"
)

print(
    "7. Summary"
)

print()

print("=" * 80)
print("DONE")
print("=" * 80)
